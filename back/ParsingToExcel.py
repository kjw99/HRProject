# /// script
# requires-python = ">=3.10"
# dependencies = [
#      "pandas",
#      "openpyxl",
#      "pymupdf",
#      "python-docx",
#      "pyhwp",
#      "python-dotenv",
# ]
# ///

import math
import os
import unicodedata
from typing import Any
import re


def _nfc(s: str | None) -> str:
    """한글 등 조합 문자(NFD) 처리·호환 문자 정돈 후 반환 DOCX/HWP/OS마다 깨져 보임 방지."""
    if s is None:
        return ""
    return unicodedata.normalize("NFC", str(s)).strip()
import subprocess
import glob
import json
import urllib.request
import urllib.error
import fitz
import pandas as pd

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILENAME = "지원자_통합관리.xlsx"
OUTPUT_PATH = os.path.join(BASE_DIR, OUTPUT_FILENAME)
if load_dotenv:
    load_dotenv(os.path.join(BASE_DIR, ".env"))

# 파싱 결과·API·엑셀 공통 필드명 (프론트 ApplicantInfo JSON 키와 동일)
K_ID = "id"
K_NAME = "name"
K_BIRTH_DATE = "birthDate"
K_CONTACT = "contact"
K_EMAIL = "email"
K_ADDR = "address"
K_ORIGINAL_JOB_ROLE = "originalJobRole"
K_FINAL_EDU = "finalEducation"
K_CAREER_PD = "careerPeriod"
K_CAREER_CO = "careerCompany"
K_CAREER_ROLE = "careerRole"
K_FILE_TYPE = "fileType"
K_EXTRACT_CONF = "extractionConfidence"

# DB/ERD 보조 필드 (API·persist 공통 — ApplicantInfo 확장·엑셀에는 미포함 가능)
K_GENDER = "gender"
K_VETERAN_ELIGIBILITY = "veteranEligibility"
K_DISABILITY = "disability"
K_DESIRED_LOCATION = "desiredLocation"
K_DESIRED_SALARY = "desiredSalary"

K_EDUCATION_ROWS = "educationRows"
K_MILITARY_ROWS = "militaryRows"
K_QUALIFICATION_ROWS = "qualificationRows"
K_EXPERIENCE_ROWS = "experienceRows"
K_STATEMENT_ROWS = "statementRows"

# 병역 등 추출 결과가 비었을 때(신뢰할 수 없는 깨짐 포함) 사용자에게 안내하기 위한 문구
NOT_EXTRACTABLE_PLACEHOLDER = "담당자 확인 필요"

# 학력: '대학원'만 있어 석사/박사 구분 불가일 때 completion_status에 넣는 안내
EDU_LEVEL_GRAD_UNCLEAR_PLACEHOLDER = "석사인지 박사인지 확인 필요"

_STRUCTURE_RESULT_KEYS = frozenset(
    {
        K_EDUCATION_ROWS,
        K_MILITARY_ROWS,
        K_QUALIFICATION_ROWS,
        K_EXPERIENCE_ROWS,
        K_STATEMENT_ROWS,
    }
)
_SUPPLEMENT_SCALAR_KEYS = frozenset(
    {
        K_GENDER,
        K_VETERAN_ELIGIBILITY,
        K_DISABILITY,
        K_DESIRED_LOCATION,
        K_DESIRED_SALARY,
    }
)

# ─── 공통 정규식 ────────────────────────────────────────────────────────────
# 날짜 범위: 2020.03~현재 / 2020.03-2024.01 / 2020.03 ~ 2024.01 등
_DATE_RANGE = re.compile(
    r"(\d{4}[.\-/]\d{1,2}"
    r"\s*[~\-–—]\s*"
    r"(?:\d{4}[.\-/]\d{1,2}|현재|재직|재학|중|진행))"
)
# 단독 연월: 2020.03 / 2020-03
_YEARMONTH = re.compile(r"\d{4}[.\-]\d{1,2}")
# 학교명
_SCHOOL = re.compile(r"([가-힣]{2,}\s*(?:대학교|대학원|대학))")
# 전공(복수 전공명 구분: 온점·가운뎃점·middot·bullet·전각점 등 — 순수 [가-힣]만 쓰면 '컴퓨터·전자…' 앞 접두가 잘림)
_MAJOR_SYM = r"[가-힣.·∙‧・•․．。･]"
_MAJOR_PATS = [
    re.compile(
        rf"({_MAJOR_SYM}{{2,}}\s*(?:학과|학부|공학과|공학부|전공|계열))"
    ),
    re.compile(
        rf"({_MAJOR_SYM}{{2,}}(?:공학|과학|시스템|전자|기계|경영|경제|컴퓨터|소프트웨어|정보|통신|디자인|미디어|통계))"
    ),
]
_GRAD_STATUS = ["졸업", "재학", "수료", "중퇴", "편입", "예정"]


def infer_education_level(
    school_name: str | None,
    department: str | None,
    level_cell: str | None,
    row_text: str | None,
) -> str | None:
    """
    educations.completion_status 저장용 최종학력 구분.
    반환: 고등학교 | 학사 | 석사 | 박사 | '대학원'만 명시된 경우 EDU_LEVEL_GRAD_UNCLEAR_PLACEHOLDER,
    그 외 판단 불가 시 None.
    """
    parts = [level_cell, school_name, department, row_text]
    full = " ".join((p or "").strip() for p in parts if (p or "").strip())
    if not full:
        return None
    ns = re.sub(r"\s+", "", full)
    low = full.casefold()
    sn = re.sub(r"\s+", "", school_name or "")

    if "중학교" not in ns and "초등학교" not in ns:
        if "고등학교" in ns:
            return "고등학교"
        if "고교" in ns:
            return "고등학교"

    if re.search(r"\bph\.?\s*d\.?|doctoral|\bdoctor\s+of\b", low) or "박사" in ns:
        return "박사"
    if "석사" in ns:
        return "석사"
    if re.search(r"\b(?:master(?:'s)?|mba)\b", low):
        return "석사"

    if "대학원" in ns:
        return EDU_LEVEL_GRAD_UNCLEAR_PLACEHOLDER

    if "전문대학" in ns or "전문대" in ns:
        return "학사"
    if "대학교" in ns:
        return "학사"
    if sn and (sn.endswith("대학교") or (sn.endswith("대학") and not sn.endswith("대학원"))):
        return "학사"
    if re.search(r"[가-힣]{2,}대학\s*\(", full):
        return "학사"

    return None


_EDU_SKIP = {"학력사항", "학력구분", "학교명", "전공", "재학기간", "소재지", "학점",
             "졸업구분", "입학", "구분", "기간", "학교", "학과명", "전공 및 학위구분", "-",
             "단과대학", "단과대학(학부)", "학부", "학 과", "평균학점", "연봉", "직급",
             "부서", "담당업무", "회사명", "기 간", "년", "월", "성명", "생년월일"}
_CAREER_KWS = [
    "MCU", "펌웨어", "개발", "사원", "대리", "과장", "차장", "부장", "팀장",
    "엔지니어", "Engineer", "연구", "설계", "제어", "테크", "솔루션",
    "오토", "센서", "R&D", "시스템", "소프트웨어", "하드웨어", "매니저",
    "직원", "재직", "근무", "입사", "퇴사", "(주)", "㈜", "주식회사",
    "회사", "기업", "산업", "인턴", "계약직", "정규직", "선임", "수석", "책임",
]
_EDU_CTX_KWS = ["대학교", "대학원", "대학", "고등학교", "수료", "양성과정", "사관학교", "중학교"]
_NOISE_LINE_KWS = {
    "학력사항", "학력구분", "학교명", "단과대학", "단과대학(학부)", "학 과", "평균학점", "소재지",
    "경력사항", "회사명", "직 급", "부 서", "담당업무", "연 봉", "해외경험", "외국어",
    "자격사항", "병역사항", "자기소개", "기 간", "기간", "년", "월", "구분", "평가년월",
    "활 동  내 역", "자격종류", "등급", "기 관", "국가보훈", "대상여부",
}


def _parse_edu_from_text(text: str, all_dates: list[str]) -> tuple[str, str]:
    """텍스트에서 최종학력과 학력날짜 추출 (표 없는 포맷 대응 폴백)"""
    best_edu = ""
    best_date = ""
    best_score = -1
    # 학교명을 손상시키지 않는 멀티워드 헤더만 제거 ('학교' 단독 단어는 제외)
    _safe_skip = ["학력사항", "학력구분", "학교명", "재학기간", "졸업구분", "학점",
                  "소재지", "학력사항", "전공 및 학위구분"]
    lines = text.split("\n")

    for i, line in enumerate(lines):
        if not any(kw in line for kw in ["대학교", "대학원", "대학"]):
            continue
        temp = line
        for h in _safe_skip:
            temp = temp.replace(h, "")
        m = _SCHOOL.search(temp)
        if not m:
            continue
        school_name = m.group(1).replace(" ", "")
        if school_name in _EDU_SKIP or "단과대학" in school_name:
            continue
        ctx = " ".join(lines[max(0, i - 2): i + 4])

        major_name = ""
        remaining = temp[m.end():] or ctx
        for mp in _MAJOR_PATS:
            mm = mp.search(remaining)
            if mm:
                major_name = mm.group(1).strip()
                break

        status = next((gs for gs in _GRAD_STATUS if gs in ctx), "")

        edu_date = ""
        school_pos = text.find(school_name)
        best_dist = 10**9
        for d in all_dates:
            d_pos = text.find(d)
            if d_pos < 0:
                continue
            if d in ctx or abs(d_pos - school_pos) < 250:
                dist = abs(d_pos - school_pos)
                if dist < best_dist:
                    best_dist = dist
                    edu_date = d

        # 최종학력은 "학교명 + 기간"이 함께 있는 항목만 채택
        if not edu_date:
            continue

        score = 0
        if "대학원" in school_name:
            score += 1
        elif "대학교" in school_name or school_name.endswith("대학"):
            score += 2
        if major_name:
            score += 1
        if status:
            score += 1

        candidate = _build_edu_str(school_name, major_name, status, edu_date)
        if score > best_score:
            best_score = score
            best_edu = candidate
            best_date = edu_date
    return best_edu, best_date


def _parse_career_from_text(
    text: str, all_dates: list[str], edu_date: str
) -> tuple[str, str, str]:
    """텍스트에서 경력기간/회사/직무 추출 (표 없는 포맷 폴백)
    반환: (경력기간, 경력회사, 경력직무)  각각 \\n 구분
    """
    entries: list[dict] = []
    seen: set[str] = set()

    for d in all_dates:
        if d == edu_date or d in seen:
            continue
        pos = text.find(d)
        ctx = text[max(0, pos - 200): pos + 200]

        if not any(ckw.upper() in ctx.upper() for ckw in _CAREER_KWS):
            continue
        if any(ekw in ctx for ekw in _EDU_CTX_KWS):
            continue

        seen.add(d)
        company, job = "", ""

        # 날짜가 포함된 줄에서 회사명/직무 추출
        for line in ctx.split("\n"):
            if d not in line:
                continue
            line = line.strip()
            date_pos = line.find(d)
            before = line[:date_pos].strip()
            after = line[date_pos + len(d):].strip()
            combined = (before + " " + after).strip()

            # 회사명: 첫 한글 단어 그룹 (일반적으로 줄 앞쪽에 위치)
            comp_m = re.search(r"([가-힣]{2,}[\w가-힣]*)", combined)
            if comp_m:
                company = comp_m.group(1)

            # 직무: 직무 관련 키워드를 포함하는 한글 구문
            job_m = re.search(
                r"([가-힣/]{2,}(?:개발|분석|관리|기획|설계|마케팅|영업|연구|사원|대리|과장|팀장|엔지니어))",
                combined,
            )
            if job_m:
                job = job_m.group(1)
            break

        entries.append({"period": d, "company": company, "job": job})

    return (
        "\n".join(e["period"]  for e in entries),
        "\n".join(e["company"] for e in entries),
        "\n".join(e["job"]     for e in entries),
    )


def _extract_basic(text: str) -> dict:
    """텍스트에서 기본 정보 추출 (이름/생년/연락처/이메일/주소/지원직무)"""
    name_m = re.search(
        r"(?:성\s*명|이\s*름|이름|Name)\s*(?:\([^)]*\))?\s*[:：]?\s*([가-힣]{2,5})", text
    )
    birth_m = re.search(r"(\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2})", text)
    phone_m = re.search(
        r"((?:010|011|016|017|018|019)[-\s]?\d{3,4}[-\s]?\d{4})",
        text,
    ) or re.search(
        r"((?:02|031|032|033|041|042|043|044|051|052|053|054|055|061|062|063|064|070)"
        r"[-\s]?\d{3,4}[-\s]?\d{4})",
        text,
    )
    # 로컬파트에 한글 등(IDN·국내 양식) 허용 · \w 는 유니코드 시 한글 포함
    email_m = re.search(
        r"(?u)([\w\-+.%가-힣]+@[a-zA-Z0-9](?:[a-zA-Z0-9.-]*[a-zA-Z0-9])?\.[a-zA-Z]{2,})",
        text,
    )
    addr_m = re.search(r"(?:주\s*소|거\s*주\s*지|Address)\s*[:：]?\s*([^\n]{5,})", text)
    job_m = re.search(
        r"(?:지\s*원\s*직\s*무|지\s*원\s*분\s*야|직\s*무|Position|지원직무)\s*[:：]?\s*"
        r"([가-힣\s\w/.\(\)]+?)(?=\n|경력여부|희망연봉|희망|About|성명|생년|$)",
        text,
    )
    return {
        K_NAME: name_m.group(1).strip() if name_m else "",
        K_BIRTH_DATE: birth_m.group(1) if birth_m else "",
        K_CONTACT: phone_m.group(1).strip() if phone_m else "",
        K_EMAIL: email_m.group(1) if email_m else "",
        K_ADDR: addr_m.group(1).strip() if addr_m else "",
        K_ORIGINAL_JOB_ROLE: job_m.group(1).strip() if job_m else "",
    }


def _normalize_hwp_table_text(text: str) -> str:
    """HWP 표가 줄 단위로 분해된 경우를 날짜 중심으로 정규화"""
    normalized = text
    # 이메일 분절: 한글/영문 로컬파트 모두 co\nm -> com 복원
    normalized = re.sub(
        r"([a-zA-Z0-9._%+\-가-힣]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2})\s*\n\s*([a-zA-Z]{1,3})",
        r"\1\2",
        normalized,
    )
    # 2026 \n 04 -> 2026.04
    normalized = re.sub(r"(\d{4})\s*\n\s*(\d{1,2})(?=\s*(?:\n|$))", r"\1.\2", normalized)
    # 2026.04 \n - \n 2026.07 -> 2026.04~2026.07
    normalized = re.sub(
        r"(\d{4}[.\-]\d{1,2})\s*\n\s*[-~]\s*\n\s*(\d{4}[.\-]\d{1,2})",
        r"\1~\2",
        normalized,
    )
    return normalized


def _normalize_pdf_text(text: str) -> str:
    """PDF 줄바꿈 분절 텍스트를 파싱 친화 형태로 정규화"""
    normalized = text
    # 이메일 분절: 한글/영문 로컬파트 모두 co\nm -> com 복원
    normalized = re.sub(
        r"([a-zA-Z0-9._%+\-가-힣]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2})\s*\n\s*([a-zA-Z]{1,3})",
        r"\1\2",
        normalized,
    )
    # 휴대폰 분절: 010-1234-1\n234 -> 010-1234-1234
    normalized = re.sub(
        r"((?:010|011|016|017|018|019)[-\s]?\d{3,4}[-\s]?\d{1,2})\s*\n\s*(\d{2,4})",
        r"\1\2",
        normalized,
    )
    # 연월 분절: 2026 04 -> 2026.04
    normalized = re.sub(r"(\d{4})\s+(\d{1,2})(?=\s*(?:\n|$))", r"\1.\2", normalized)
    # 기간 분절: YYYY.MM \n - \n YYYY.MM -> YYYY.MM~YYYY.MM
    normalized = re.sub(
        r"(\d{4}[.\-]\d{1,2})\s*\n\s*[-~]\s*\n\s*(\d{4}[.\-]\d{1,2})",
        r"\1~\2",
        normalized,
    )
    # 같은 줄 변형: YYYY.MM - YYYY.MM -> YYYY.MM~YYYY.MM
    normalized = re.sub(
        r"(\d{4}[.\-]\d{1,2})\s*[-~]\s*(\d{4}[.\-]\d{1,2})",
        r"\1~\2",
        normalized,
    )
    return normalized


def _parse_hwp_sections(text: str) -> tuple[str, str, str, str]:
    """HWP 표형/자유형 이력서에서 학력/경력을 줄 단위로 보강 추출"""
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if not lines:
        return "", "", "", ""

    def _find_index(keyword: str) -> int:
        for i, ln in enumerate(lines):
            if keyword in ln:
                return i
        return -1

    edu_start = _find_index("학력사항")
    career_start = _find_index("경력사항")
    career_end = _find_index("해외경험")

    def _line_has_header(line: str) -> bool:
        return any(h in line for h in _NOISE_LINE_KWS)

    # 학력: 마지막 대학(원) 행 + 인접 날짜 조합
    edu_str = ""
    edu_date = ""
    best_edu_score = -1
    edu_scan_lines = (
        lines[edu_start + 1: career_start if career_start > edu_start else len(lines)]
        if edu_start >= 0 else lines
    )
    for i, ln in enumerate(edu_scan_lines):
        if not re.fullmatch(r"[가-힣A-Za-z0-9() ]{2,}\s*(?:대학교|대학원|대학)", ln):
            continue
        if ln in _EDU_SKIP or "단과대학" in ln:
            continue

        major = ""
        for nxt in edu_scan_lines[i + 1:i + 5]:
            if _line_has_header(nxt):
                continue
            if _line_looks_like_major_field(nxt):
                major = nxt
                break

        # 주변 10줄에서 y,m,y,m 패턴 찾기
        window = " ".join(edu_scan_lines[max(0, i - 10): i + 2])
        nums = re.findall(r"\b(\d{4})[.\-]?(\d{1,2})\b", window)
        period = ""
        if len(nums) >= 2:
            s_y, s_m = nums[-2]
            e_y, e_m = nums[-1]
            period = f"{s_y}.{s_m}~{e_y}.{e_m}"

        status = next((gs for gs in _GRAD_STATUS if gs in " ".join(edu_scan_lines[i:i + 8])), "")
        # "학교명 + 기간"이 있어야 최종학력 후보로 인정
        if not period:
            continue

        school = ln.replace(" ", "")
        score = 0
        if "대학원" in school:
            score += 1
        elif "대학교" in school or school.endswith("대학"):
            score += 2
        if major:
            score += 1
        if status:
            score += 1

        if score > best_edu_score:
            best_edu_score = score
            edu_date = period
            edu_str = _build_edu_str(school, major, status, edu_date)

    # 경력: 경력 구간 내(없으면 전체 스캔) 날짜 범위 + 회사/직무 추출
    career_period = ""
    career_company = ""
    career_job = ""
    end = career_end if career_end > career_start >= 0 else len(lines)
    clines = lines[career_start + 1:end] if career_start >= 0 else lines

    ym_positions: list[tuple[int, str, str]] = []
    for idx, ln in enumerate(clines):
        m = re.fullmatch(r"(\d{4})[.\-]?(\d{1,2})", ln)
        if m:
            ym_positions.append((idx, m.group(1), m.group(2)))

    # 가능한 모든 pair 중에서 직업 문맥이 있는 첫 pair 선택
    for p in range(0, len(ym_positions) - 1, 2):
        _, s_y, s_m = ym_positions[p]
        e_idx, e_y, e_m = ym_positions[p + 1]
        tail = clines[e_idx + 1:e_idx + 12]
        tail_text = " ".join(tail)
        if not any(k in tail_text for k in _CAREER_KWS) and "회사" not in tail_text:
            continue
        career_period = f"{s_y}.{s_m}~{e_y}.{e_m}"
        filtered = [
            x for x in tail
            if x not in {"-", "회사명", "직급", "부서", "담당업무", "연봉", "년", "월"}
            and not _line_has_header(x)
        ]
        if filtered:
            career_company = filtered[0]
        # 직무는 키워드 우선
        career_job = next(
            (
                x for x in filtered[1:]
                if re.search(r"(개발|분석|관리|기획|설계|연구|엔지니어|담당|업무|부서|직무|QA|테스트)", x, re.I)
            ),
            filtered[1] if len(filtered) > 1 else "",
        )
        break

    return edu_str, career_period, career_company, career_job


def _parse_pdf_sections(text: str) -> tuple[str, str, str, str]:
    """PDF 표형 이력서의 학력/경력 보강 추출"""
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if not lines:
        return "", "", "", ""

    def _find_index(keyword: str) -> int:
        for i, ln in enumerate(lines):
            if keyword in ln:
                return i
        return -1

    def _is_header(line: str) -> bool:
        return any(h in line for h in _NOISE_LINE_KWS)

    edu_start = _find_index("학력사항")
    career_start = _find_index("경력사항")
    career_end = _find_index("해외경험")

    # 학력
    edu_str = ""
    best_score = -1
    edu_date = ""
    edu_lines = (
        lines[edu_start + 1: career_start if career_start > edu_start else len(lines)]
        if edu_start >= 0 else lines
    )
    for i, ln in enumerate(edu_lines):
        if not re.search(r"(대학교|대학원|대학)$", ln):
            continue
        if _is_header(ln) or ln in _EDU_SKIP:
            continue

        major = ""
        for nxt in edu_lines[i + 1:i + 5]:
            if _is_header(nxt):
                continue
            if _line_looks_like_major_field(nxt):
                major = nxt
                break

        window = " ".join(edu_lines[max(0, i - 10): i + 2])
        ranges = _DATE_RANGE.findall(window)
        period = ranges[-1] if ranges else ""
        if not period:
            yms = _YEARMONTH.findall(window)
            if len(yms) >= 2:
                period = f"{yms[-2]}~{yms[-1]}"
        if not period:
            continue

        status = next((gs for gs in _GRAD_STATUS if gs in " ".join(edu_lines[i:i + 8])), "")
        school = ln.replace(" ", "")
        score = _score_edu_candidate(school, major, status, period)
        if score > best_score:
            best_score = score
            edu_date = period
            edu_str = _build_edu_str(school, major, status, edu_date)

    # 경력
    career_period = ""
    career_company = ""
    career_job = ""
    end = career_end if career_end > career_start >= 0 else len(lines)
    clines = lines[career_start + 1:end] if career_start >= 0 else lines

    for i, ln in enumerate(clines):
        if not _DATE_RANGE.search(ln):
            continue
        tail = clines[i + 1:i + 12]
        tail_text = " ".join(tail)
        if "회사" not in tail_text and not any(k in tail_text for k in _CAREER_KWS):
            continue
        career_period = _DATE_RANGE.search(ln).group(1)
        filtered = [
            x for x in tail
            if not _is_header(x) and x not in {"-", "/", "`8000"} and not _DATE_RANGE.search(x)
        ]
        if filtered:
            career_company = filtered[0]
        career_job = next(
            (
                x for x in filtered[1:]
                if re.search(r"(개발|분석|관리|기획|설계|연구|엔지니어|담당|업무|부서|직무|QA|테스트)", x, re.I)
            ),
            filtered[1] if len(filtered) > 1 else "",
        )
        break

    return edu_str, career_period, career_company, career_job


def _find_col(header_row: list[str], keywords: list[str]) -> int:
    """헤더 행에서 키워드가 포함된 첫 컬럼 (-1 = 없음). 대소문자 무관(certificate 등)."""
    for j, cell in enumerate(header_row):
        cl = cell.strip().lower()
        for kw in keywords:
            if kw.strip().lower() in cl:
                return j
    return -1


_CELL_YYYY_MM_ONLY = re.compile(r"^\s*(20\d{2})\.(\d{1,2})\s*$")


def _cell_is_yyyy_mm_only(cell: str) -> bool:
    """첫 열이 `2026.01`처럼 월 단위만 있는 형태(헤더 없는 자격/어학 표용)."""
    return bool(_CELL_YYYY_MM_ONLY.match((cell or "").strip()))


def _headerless_qual_match_count(rows: list[list[str]]) -> int:
    """각 행이 [YYYY.MM | 자격/어학명 | …] 패턴과 맞는지 개수 세기."""
    n = 0
    for row in rows:
        if len(row) < 2:
            continue
        c0 = (row[0] or "").strip()
        c1 = (_nfc(row[1] or "").strip())
        if not _cell_is_yyyy_mm_only(c0) or not c1:
            continue
        if _qual_line_is_noise(c1):
            continue
        n += 1
    return n


def _is_headerless_qual_grid(rows: list[list[str]]) -> bool:
    """헤더 행 없이 날짜 열만 있는 자격 표 (첫 줄부터 데이터)."""
    if len(rows) < 2:
        return False
    return _headerless_qual_match_count(rows) >= 2


def _build_edu_str(school: str, major: str, status: str, period: str) -> str:
    parts = [p for p in [school, major, status] if p and p.strip() and p not in _EDU_SKIP]
    if period:
        parts.append(f"({period})")
    return " ".join(parts)


def _line_looks_like_major_field(line: str) -> bool:
    """
    표에서 '전공/학과' 열이 '한글과', '경영학과'처럼 나오는 형태.
    '한글과'는 '학과' 부분문자열이 없어 기존(if '학과' in line)만으로는 누락됨.
    """
    t = _nfc(line).strip()
    if len(t) < 2 or len(t) > 48:
        return False
    if any(k in t for k in ("학과", "학부", "전공", "계열")):
        return True
    if re.match(r"^[가-힣·・]{2,34}과\s*$", t):
        return True
    return False


def _score_edu_candidate(school: str, major: str, status: str, period: str) -> int:
    """최종학력 후보 점수: 기간 필수, 학위/정보량 우선"""
    if not period:
        return -1
    score = 0
    if "대학원" in school:
        score += 1
    elif "대학교" in school or school.endswith("대학"):
        score += 2
    if major:
        score += 1
    if status:
        score += 1
    return score


def _is_bad_edu_value(value: str) -> bool:
    if not value.strip():
        return True
    if value.strip().startswith("("):
        return True
    return not bool(re.search(r"(대학교|대학원|대학)", value))


def _slice_section_text(
    text: str,
    start_keywords: tuple[str, ...],
    end_keywords: tuple[str, ...],
) -> str:
    """start_keywords 중 먼저 나오는 헤더 다음부터, end_keywords 직전까지 잘라 반환."""
    t = text.replace("\r", "")
    best = -1
    start_at = 0
    for kw in start_keywords:
        p = t.find(kw)
        if p >= 0 and (best < 0 or p < best):
            best = p
            start_at = p + len(kw)
    if best < 0:
        return ""
    rest = t[start_at:]
    end_pos = len(rest)
    for ek in end_keywords:
        if ek and ek in start_keywords:
            continue
        ep = rest.find(ek)
        if ep >= 0:
            end_pos = min(end_pos, ep)
    return rest[:end_pos].strip()


def _normalize_intro_header_spacing(text: str) -> str:
    """DOCX/PDF 표 등에서 자 기 소 개(서)·입 사 지 원 서 줄을 한 토큰으로 묶음."""
    s = text.replace("\r", "")
    # '서'가 없거나 붙어 있어도 '자 기 소 개' → 자기소개서
    s = re.sub(r"(?i)자\s*기\s*소\s*개(?:\s*서)?", "자기소개서", s)
    s = re.sub(r"입\s*사\s*지\s*원\s*서", "입사지원서", s)
    return s


def _slice_intro_text_after_line_pr_heading(text: str, end_keywords: tuple[str, ...]) -> str:
    """삼성 등 레이아웃: 단독 'PR'(Cover letter) 줄 다음부터 본문."""
    m = re.search(r"(?m)^\s*PR\s*$", text)
    if not m:
        return ""
    rest = text[m.end() :].lstrip()
    end_pos = len(rest)
    for ek in end_keywords:
        if not ek:
            continue
        ep = rest.find(ek)
        if ep >= 0:
            end_pos = min(end_pos, ep)
    return rest[:end_pos].strip()


def _slice_intro_after_numbered_samsung_heading(text: str, end_keywords: tuple[str, ...]) -> str:
    """PR 없이 번호 매겨진 자소 문항 첫 줄(예: '1. 지원동기')부터."""
    m = re.search(
        r"(?m)^\s*1\s*[.)]\s*[^\n]*(?:지원동기|지원\s*직무|입사\s*후)[^\n]*$",
        text,
    )
    if not m:
        return ""
    rest = text[m.start() :]
    ep = len(rest)
    for ek in end_keywords:
        if ek:
            p2 = rest.find(ek)
            if p2 >= 0:
                ep = min(ep, p2)
    return rest[:ep].strip()


def _slice_intro_text_longest_candidate(
    text: str,
    start_groups: tuple[tuple[str, ...], ...],
    end_keywords: tuple[str, ...],
) -> str:
    """여러 시작 마커·전략 중 가장 긴 비어있지 않은 블록을 선택."""
    best = ""
    for starts in start_groups:
        cand = (
            _slice_section_text(text, starts, end_keywords).strip()
            if starts
            else ""
        )
        if len(cand) > len(best):
            best = cand

    after_pr = _slice_intro_text_after_line_pr_heading(text, end_keywords).strip()
    if len(after_pr) > len(best):
        best = after_pr

    numbered_tail = _slice_intro_after_numbered_samsung_heading(text, end_keywords).strip()
    if len(numbered_tail) > len(best):
        best = numbered_tail

    return best


_MILITARY_TYPE_KWS = [
    "복무완료",
    "군필",
    "만기제대",
    "전역자",
    "면제",
    "미필",
    "공익근무",
    "산업기능요원",
    "전역",
    "예비역",
    "복무 중",
    "면 제",
]
_MILITARY_SVC_KWS = [
    "육군",
    "해군",
    "공군",
    "해병대",
    "의무경찰",
    "의무소방",
    "카투사",
    "상근",
]


def _likely_pdf_glyph_garbage_no_hangul(s: str) -> bool:
    """
    PDF 폰트/ToUnicode 오류 등으로 한글 대신 무의미한 한자 문자만 들어오는 패턴 배제.
    (예: '신체' 자리가 氠瑢처럼 잘못 추출되는 경우.)
    한글 음절(가–힣)이 하나도 없고, 남는 부분이 CJK 호환·통합 한자뿐이면 신뢰하지 않음.
    """
    if not s.strip():
        return False
    if re.search(r"[\uac00-\ud7af]", s):
        return False
    # 허용: 숫자·라틴·구두점·공백만이면 간단 표기 가능
    core = re.sub(r"[\s0-9A-Za-z.,\-/+()°℃·:×÷%‰]+", "", s)
    if not core.strip():
        return False
    for ch in core:
        o = ord(ch)
        if o == 0x3000:
            continue
        if 0x3400 <= o <= 0x9FFF or 0xF900 <= o <= 0xFAFF:
            continue
        return False
    return True


_EXEMPTION_REASON_JUNK_KEYS = frozenset(
    {
        "자기소개",
        "자기소개서",
        "학력사항",
        "경력사항",
        "자격사항",
        "병역사항",
        "어학",
        "외국어",
        "해외경험",
        "수상",
        "종교",
    }
)


def _strip_leading_resume_section_headers(reason: str) -> str:
    """면제사유 문자열 선두에 단독으로 붙은 섹션 제목만 제거(반복 허용)."""
    s = reason.strip()
    # 셀이 '자기소개' 하나만 들어오거나 헤더+본문 시작이 줄 때 선두만 직접 제거
    while True:
        prev = s
        s = re.sub(
            r"^(?:자기소개(?:서)?|학력사항|경력사항|자격사항|병역사항)(\s+|[:：])?\s*",
            "",
            s,
            count=1,
        ).strip()
        if s == prev:
            break

    sep = (
        r"(?:자기소개(?:서)?|학력사항|경력사항|자격사항|병역사항)"
    )
    m = re.search(r"\s+(?:" + sep + r")(?:\s+|$).*", s, flags=re.DOTALL)
    if m:
        s = s[: m.start()].strip()
    return s.strip()


def _clean_military_exemption_reason(raw: str) -> str | None:
    """
    병역 블록에서 '면제' 뒤로 잡힌 문자열 정리.
    표에서 '면제' 옆 칸에 '종교'(신앙 구분 헤더)만 붙은 경우 면제사유로 넣지 않음.
    '종교 기독교'처럼 헤더+값이 붙은 경우 실제 값만 남김.
    같은 줄로 합쳐진 '자기소개' 등 다음 섹션 헤더는 제거한다.
    PDF 추출 깨짐(한자만·한글 없음)은 저장하지 않음.
    """
    if not raw:
        return None
    s = raw.strip()
    if not s:
        return None
    s = _strip_leading_resume_section_headers(s)
    if not s:
        return None
    if s in _EXEMPTION_REASON_JUNK_KEYS:
        return None
    if _likely_pdf_glyph_garbage_no_hangul(s):
        return None
    # 단독 '종교'는 보통 표 헤더가 줄과 합쳐진 경우가 많음
    if s == "종교":
        return None
    # '종교 기독교'처럼 헤더+신앙만 (공백 구분); '종교적 양심'은 건드리지 않음
    if s.startswith("종교 "):
        tail = s[3:].strip()
        return tail[:500] if tail else None
    return s[:500]


_EXEMPTION_TRUST_PATTERN = re.compile(
    r"""(
        신체|등위|등\s*위\s*판정|등급\s*판정|판정\s*[:]?\s*|
        질병|장애|중증|처분|병증|증명(?:서|원)?|면제증|사유서|판정등급|면제판정|
        외국|영주국|예술|예비역|처분번호|복무\s*판정|
        종교|양심|선원|철회|
        [0-9]{1,2}\s*[급]|제\s*[0-9]+\s*[급]
    )""",
    re.VERBOSE | re.UNICODE,
)


def exemption_reason_trusted_enough(reason: str) -> bool:
    """
    조금이라도 면제사유로 보기 애매하면 False → 호출처에서 담당자 확인 플레이스홀더로 교체.

    간단 패턴 또는 충분한 길이의 줄글 한글 등으로만 신뢰.
    """
    t = (reason or "").strip()
    if not t or t == NOT_EXTRACTABLE_PLACEHOLDER:
        return False
    if _likely_pdf_glyph_garbage_no_hangul(t):
        return False
    if len(t) == 1:
        return False
    if _EXEMPTION_TRUST_PATTERN.search(t):
        return True
    hangul_chars = sum(1 for ch in t if "\uac00" <= ch <= "\ud7af")
    # 패턴에 안 걸리는 긴 줄글 한글 한정(가짜 열 문자열 줄이기 위해 기준 김)
    if len(t) >= 48 and hangul_chars >= 24:
        return True
    if len(t) >= 72 and hangul_chars >= 14:
        return True
    return False


def _empty_strip(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _parse_period_to_raw_pair(period: str) -> tuple[str | None, str | None]:
    """재직/복무 기간 문자열에서 시작·끝 날짜 후보 문자열."""
    if not period or not period.strip():
        return None, None
    period = period.strip()
    dr = _DATE_RANGE.search(period)
    if dr:
        chunk = dr.group(1)
        parts = re.split(r"\s*[~\-–—]\s*", chunk, maxsplit=1)
        left = parts[0].strip() if parts else ""
        right = parts[1].strip() if len(parts) > 1 else ""
        if right in ("현재", "재직", "재학", "중", "진행"):
            right = ""
        return left or None, right or None
    yms = _YEARMONTH.findall(period)
    if len(yms) >= 2:
        return yms[0], yms[-1]
    if len(yms) == 1:
        return yms[0], None

    kms = list(re.finditer(r"(\d{4})\s*년\s*(\d{1,2})\s*월", period))
    if len(kms) >= 2:
        ma, mb = kms[0], kms[-1]
        return (
            f"{ma.group(1)}.{int(ma.group(2)):02d}",
            f"{mb.group(1)}.{int(mb.group(2)):02d}",
        )
    if len(kms) == 1:
        ma = kms[0]
        return (f"{ma.group(1)}.{int(ma.group(2)):02d}", None)

    return None, None


def parse_education_period_raw_pair(
    period_cell: str,
    row_text: str | None,
) -> tuple[str | None, str | None]:
    """
    학력 행 재학기간: 셀만 비거나 형식 특이하면 같은 행 전체 문자열에서 보완 추출.

    다른 모듈(저장소 폴백)에서도 사용 가능하게 공개.
    """
    p = (period_cell or "").strip()
    rtx = (row_text or "").strip()
    if p:
        a0, a1 = _parse_period_to_raw_pair(p)
        if a0 or a1:
            return a0, a1
    if rtx:
        return _parse_period_to_raw_pair(rtx)
    return None, None


def split_merged_education_text(blob: str) -> dict[str, str | None]:
    """
    한 줄·한 셀에 '학교 전공 졸업 (2022.03~2026.02)'처럼 붙은 문자열을
    학교명 / 전공(또는 잔여) / 최종학력( infer_education_level ) / 재학 기간으로 분리.
    표 컬럼이 없거나 한 컬럼만 있을 때 DB educations 매핑용.
    """
    blob = _nfc(blob)
    out: dict[str, str | None] = {}
    if not blob or blob == "기타":
        return out

    period_raw = ""
    dr = _DATE_RANGE.search(blob)
    if dr:
        period_raw = dr.group(1)

    work = blob
    if dr:
        full = dr.group(0)
        wrapped = re.search(r"\(\s*" + re.escape(full) + r"\s*\)", blob)
        if wrapped:
            work = blob.replace(wrapped.group(0), " ", 1)
        else:
            work = blob.replace(full, " ", 1)
    work = re.sub(
        r"\(\s*[0-9\s.\-~/〜～∼]+\s*[~\-–—]\s*[0-9\s.\-~/〜～∼]+\s*\)",
        " ",
        work,
        count=1,
    )
    work = re.sub(r"\([^)]*[~〜～∼－][^)]*\)", " ", work, count=1)
    work = re.sub(r"\s+", " ", work).strip()

    completion = ""
    for gs in sorted(_GRAD_STATUS, key=len, reverse=True):
        if gs in work:
            completion = gs
            work = work.replace(gs, " ", 1)
            break
    work = re.sub(r"\s+", " ", work).strip()

    school = ""
    department = ""
    sm = _SCHOOL.search(work)
    if sm:
        school = sm.group(1).replace(" ", "")
        department = work[sm.end() :].strip()
    else:
        first_space = blob.find(" ")
        school = (blob[: first_space].strip() if first_space > 0 else blob)[:255]
        department = ""

    if not department:
        for mp in _MAJOR_PATS:
            mm = mp.search(work)
            if mm:
                department = mm.group(1).strip()
                break

    dept_clean = department[:500].strip() if department else ""

    a0, a1 = parse_education_period_raw_pair(period_raw, blob)

    sn_out = _nfc(school.strip()[:500] if school.strip() else "")
    dp_out = _nfc(dept_clean)
    out["school_name"] = sn_out or None
    out["department"] = dp_out or None
    out["completion_status"] = infer_education_level(
        out["school_name"],
        out["department"],
        None,
        blob,
    )
    out["period_raw"] = period_raw or None
    out["attendance_start_period_raw"] = a0
    out["attendance_end_period_raw"] = a1
    return out


def _merged_edu_blob_has_department(blob: str | None) -> bool:
    """`split_merged_education_text` 기준으로 전공/학과 문자열이 분리되는지."""
    t = _nfc(blob or "").strip()
    if not t or t == "기타":
        return False
    try:
        sp = split_merged_education_text(t)
    except Exception:
        return False
    d = sp.get("department")
    return bool(d and str(d).strip())


def _refine_education_row_dict(row: dict[str, Any]) -> dict[str, Any]:
    """표 인식 결과 한 덩어리 셀을 합산된 경우 학교명·전공·기간 보정."""
    sn = _nfc(row.get("school_name"))
    mj = _nfc(row.get("department"))
    ps = _nfc(row.get("period_raw"))
    combo = " ".join(x.strip() for x in (sn, mj, ps) if x and x.strip())

    messy = False
    if "~" in str(sn) or ("(" in str(sn) and ")" in str(sn)):
        messy = True
    if ("졸업" in str(sn) or "재학" in str(sn)) and len(str(sn)) > 20:
        messy = True
    if len(str(sn)) > 55:
        messy = True

    if not messy:
        return row

    try:
        sp = split_merged_education_text(combo or str(sn))
    except Exception:
        return row
    if not sp:
        return row

    row["school_name"] = sp.get("school_name") or row.get("school_name")
    row["department"] = sp.get("department") or row.get("department")
    row["completion_status"] = infer_education_level(
        row.get("school_name"),
        row.get("department"),
        None,
        combo or str(sn),
    )
    row["period_raw"] = sp.get("period_raw") or row.get("period_raw")
    if sp.get("attendance_start_period_raw"):
        row["attendance_start_period_raw"] = sp["attendance_start_period_raw"]
    if sp.get("attendance_end_period_raw"):
        row["attendance_end_period_raw"] = sp["attendance_end_period_raw"]
    return row


def _parse_military_from_block(block: str) -> list[dict]:
    if not block or len(block.strip()) < 2:
        return []
    b = block.strip()
    mtype = ""
    for kw in _MILITARY_TYPE_KWS:
        if kw in b.replace(" ", ""):
            mtype = kw
            break
    svc = ""
    for kw in _MILITARY_SVC_KWS:
        if kw in b:
            svc = kw
            break
    s0, s1 = _parse_period_to_raw_pair(b)
    ex_reason = ""
    if "면제" in b or "면 제" in b:
        # 우선 면제사유 또는 괄호 안만 쓰면 헤더(종교 등) 침범 감소
        lm = re.search(
            r"면제\s*사유\s*[\(:：]?\s*([^\n|]+)",
            b,
        )
        if not lm:
            lm = re.search(r"면제\s*\(\s*([^)\n]+)\s*\)", b)
        if not lm:
            lm = re.search(r"면제\s*[\(:：]?\s*([^\n]+)", b)
        if lm:
            ex_reason = _clean_military_exemption_reason(lm.group(1)) or ""
    exemption_context = "면제" in b.replace(" ", "") or "면 제" in b
    er = (ex_reason or "").strip()
    if (
        exemption_context
        and er
        and er != NOT_EXTRACTABLE_PLACEHOLDER
        and not exemption_reason_trusted_enough(er)
    ):
        ex_reason = NOT_EXTRACTABLE_PLACEHOLDER
    else:
        ex_reason = er
    # 면제인데 사유 문자열을 끝내 못 얻은 경우(깨짐 제거·비어 있음)
    if exemption_context and not (ex_reason or "").strip():
        ex_reason = NOT_EXTRACTABLE_PLACEHOLDER
    row: dict[str, str | None] = {
        "military_type": mtype or None,
        "military_service": svc or None,
        "military_start_period_raw": s0,
        "military_end_period_raw": s1,
        "military_rank": None,
        "exemption_reason": ex_reason or None,
    }
    if any(v for v in row.values() if v):
        return [row]
    return []


def _military_slash_segments_to_row(segment: str) -> dict[str, Any] | None:
    """예: 필 / 2012.05 ~ 2014.02 / 육군 / 병장."""
    seg = _nfc(segment.strip())
    if not seg or "/" not in seg:
        return None
    parts = [p.strip() for p in re.split(r"\s*/\s*", seg) if p.strip()]
    if len(parts) < 3:
        return None
    raw_t = parts[0]
    period_chunk = ""
    svc = None
    rank = None
    if len(parts) >= 4:
        period_chunk = parts[1]
        svc = parts[2] if len(parts) >= 3 else None
        rank = parts[3] if len(parts) >= 4 else None
    elif len(parts) == 3:
        if re.search(r"\d{4}", parts[1]):
            period_chunk, svc = parts[1], parts[2]
        else:
            return None

    mt = ""
    rt = raw_t.replace(" ", "")
    if rt in ("필",):
        mt = "군필"
    elif rt in ("면제", "미필", "예비역", "복무중", "복무중"):
        for k in ("면제", "미필", "예비역"):
            if k in rt:
                mt = k
                break
    else:
        for kw in _MILITARY_TYPE_KWS:
            if kw.replace(" ", "") in rt:
                mt = kw
                break
    if not mt and raw_t.strip():
        mt = raw_t.strip()[:32]

    s0, s1 = None, None
    if period_chunk:
        s0, s1 = _parse_period_to_raw_pair(period_chunk)
    svc_f = None
    if svc:
        for br in _MILITARY_SVC_KWS:
            if br in svc:
                svc_f = br
                break
        if svc_f is None and svc.strip():
            svc_f = svc.strip()[:48]

    ex = None
    if mt == "면제":
        tail = (svc or rank or "").strip()
        if tail and tail not in _EXEMPTION_REASON_JUNK_KEYS:
            ex = tail[:500]

    return {
        "military_type": mt or None,
        "military_service": svc_f,
        "military_start_period_raw": s0,
        "military_end_period_raw": s1,
        "military_rank": rank.strip()[:50] if rank else None,
        "exemption_reason": ex,
    }


def _military_dict_from_cells(cells: list[str]) -> dict[str, Any] | None:
    """인적 표 한 행 안에서 「병역」 레이블 옆 칸 또는 슬래시 줄."""
    for i, c in enumerate(cells):
        tl = (_nfc(c) or "").strip()
        tl_sp = tl.replace(" ", "")
        if not tl_sp:
            continue
        if tl_sp == "병역" and i + 1 < len(cells):
            val = (_nfc(cells[i + 1]) or "").strip()
            r = _single_cell_military_val(val)
            if r:
                return r
        if tl_sp == "병역구분" and i + 1 < len(cells):
            row_sl = "/".join((_nfc(x) or "").strip() for x in cells[i:] if (_nfc(x) or "").strip())
            rr = _military_slash_segments_to_row(row_sl)
            if rr and any(rr.values()):
                return rr
        # merged cell
        if "/" in tl and re.search(r"\d{4}", tl):
            rr = _military_slash_segments_to_row(tl)
            if rr and any(rr.get(k) for k in ("military_service", "military_start_period_raw", "military_type")):
                return rr
    return None


def _single_cell_military_val(val: str) -> dict[str, Any] | None:
    v = (_nfc(val) or "").strip()
    if not v or len(v) > 80:
        return None
    vns = v.replace(" ", "")
    mt = ""
    for kw in _MILITARY_TYPE_KWS:
        if kw.replace(" ", "") in vns:
            mt = kw
            break
    if not mt:
        if "만기" in v:
            mt = "만기제대"
        else:
            return None
    return {
        "military_type": mt,
        "military_service": None,
        "military_start_period_raw": None,
        "military_end_period_raw": None,
        "military_rank": None,
        "exemption_reason": None,
    }


def _military_rows_from_docx_tables(all_tables: list[list[list[str]]] | None) -> list[dict]:
    """DOCX 인적·병역 표에서 병역 한 줄 추출."""
    if not all_tables:
        return []
    out: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()

    def _sig(d: dict[str, Any]) -> tuple[Any, ...]:
        return (
            d.get("military_type"),
            d.get("military_service"),
            d.get("military_start_period_raw"),
            d.get("military_end_period_raw"),
        )

    def _try_add(d: dict[str, Any] | None) -> None:
        if not d or not any(v for v in d.values() if v):
            return
        sg = _sig(d)
        if sg in seen:
            return
        seen.add(sg)
        out.append(d)

    for rows in all_tables:
        if not rows or len(rows) < 1:
            continue
        header = [_nfc(c or "").strip() for c in rows[0]]
        hdr_one = "".join(header).replace(" ", "")
        has_mil_hdr = "병역" in hdr_one or "군별" in hdr_one or "계급" in hdr_one
        for row in rows[1:] if has_mil_hdr and len(rows) > 1 else rows:
            cells = [_nfc(c or "").strip() for c in row]
            line = "|".join(cells)
            jl = line.replace(" ", "")
            if has_mil_hdr and "/" in line and re.search(r"\d{4}", line):
                rj = _military_slash_segments_to_joined_row(header, cells)
                _try_add(rj)
            rr = _military_dict_from_cells(cells)
            _try_add(rr)
            for c in cells:
                if "/" in c and re.search(r"\d{4}", c) and re.search(r"육군|해군|공군|해병|군필|필", c):
                    _try_add(_military_slash_segments_to_row(c))
        if len(out) >= 15:
            break
    return out


def _military_slash_segments_to_joined_row(header_row: list[str], data_cells: list[str]) -> dict[str, Any] | None:
    """병역구분 / 복무기간 / 군별 / 계급 형태 표."""
    h = [_nfc(x or "").strip() for x in header_row]
    d = [_nfc(x or "").strip() for x in data_cells]
    if len(d) < 2:
        return None
    hi = {h[i].replace(" ", ""): i for i in range(len(h)) if h[i].strip()}
    parts: list[str] = []
    for label, key in (
        ("병역구분", "type"),
        ("복무기간", "period"),
        ("군별", "svc"),
        ("계급", "rank"),
    ):
        idx = -1
        for j, hl in enumerate(h):
            if label in hl.replace(" ", ""):
                idx = j
                break
        if idx >= 0 and idx < len(d) and d[idx].strip():
            parts.append(d[idx].strip())
    if len(parts) >= 2:
        line = " / ".join(parts)
        return _military_slash_segments_to_row(line.replace(" / ", "/"))
    return None


def _military_rows_from_fulltext_patterns(text: str) -> list[dict]:
    """평문에서 «병역사항 | 필 / … / 육군 / …» 등 슬래시 형식."""
    blob = _nfc(text)
    out: list[dict] = []
    seen: set[tuple[Any, ...]] = set()

    def _add(d: dict[str, Any] | None) -> None:
        if not d or not any(d.values()):
            return
        k = (
            d.get("military_type"),
            d.get("military_service"),
            d.get("military_start_period_raw"),
            d.get("military_end_period_raw"),
        )
        if k in seen:
            return
        seen.add(k)
        out.append(d)

    for m in re.finditer(
        r"(?m)[^\n]{0,48}병역(?:사항)?\s*[|｜]\s*([^\n]{3,420})",
        blob,
    ):
        chunk = m.group(1).strip()
        if "/" in chunk:
            rr = _military_slash_segments_to_row(chunk.split("\n")[0])
            _add(rr)
        else:
            _add(_single_cell_military_val(chunk.split("|")[0].strip()))

    for m in re.finditer(r"(?<![가-힣])복무\s*기간\s*[:：]?\s*([^\n]{6,240})", blob):
        s0, s1 = _parse_period_to_raw_pair(m.group(1))
        if s0 or s1:
            _add(
                {
                    "military_type": "군필",
                    "military_service": None,
                    "military_start_period_raw": s0,
                    "military_end_period_raw": s1,
                    "military_rank": None,
                    "exemption_reason": None,
                }
            )

    if len(out) > 15:
        return out[:15]
    return out


def _merge_military_extractions(
    plain_block_rows: list[dict],
    table_rows: list[dict],
    fulltext_rows: list[dict],
) -> list[dict]:
    """표 → 전체 텍스트 패턴 → 병역사항 블록 순으로 합치되, 동일 서명은 한 번만."""
    merged: list[dict] = []
    seen = set()

    def _sig(d: dict) -> tuple[Any, ...]:
        return (
            d.get("military_type"),
            d.get("military_service"),
            d.get("military_start_period_raw"),
            d.get("military_end_period_raw"),
        )

    for cand in (*table_rows, *fulltext_rows, *(plain_block_rows or [])):
        if not isinstance(cand, dict) or not any(cand.values()):
            continue
        sg = _sig(cand)
        if sg in seen:
            continue
        seen.add(sg)
        merged.append(dict(cand))

    merged.sort(
        key=lambda d: (
            bool(d.get("military_start_period_raw")),
            bool(d.get("military_service")),
            bool(d.get("military_type")),
        ),
        reverse=True,
    )
    return merged[:15]


_QUAL_NOISE_LINES = {
    "자격사항",
    "자격증",
    "종류",
    "등급",
    "발행처",
    "취득일",
    "-",
    "기 간",
    "기간",
    "활 동 내 역",
    "활 동  내 역",
    "활동내역",
    "자격종류",
    "평가년월",
    "구 분",
}

# 줄 전체가 헤더일 때 ·셀 문자 사이에 공백이 든 형태까지 공밅 제거 후 비교 (_NOISE_LINE_KWS 포함)
_QUAL_HEADER_COLLAPSED = frozenset(
    {re.sub(r"\s+", "", w) for w in (_NOISE_LINE_KWS | _QUAL_NOISE_LINES)}
    | {
        "자격증명",
        "합격번호",
        "합격증",
        "기관명",
        "발급기관",
        "언어점수",
    }
)


def _qual_line_is_noise(line: str) -> bool:
    """표 헤더·항목 이름만 한 줄로 잡혀 자격증처럼 넣히는 경우 차단."""
    t = _nfc(line)
    if not t:
        return True
    if t in _QUAL_NOISE_LINES:
        return True
    collapsed = re.sub(r"\s+", "", t)
    if collapsed in _QUAL_HEADER_COLLAPSED:
        return True
    return False


def _parse_qualification_from_block(block: str) -> list[dict]:
    rows: list[dict] = []
    for line in block.split("\n"):
        line = line.strip()
        if len(line) < 3:
            continue
        if _qual_line_is_noise(line):
            continue
        if re.fullmatch(r"[\s\-.0-9]+", line):
            continue
        dm = re.search(r"(20\d{2})[.\-/년]\s*(\d{1,2})", line)
        issue_raw = None
        if dm:
            issue_raw = f"{dm.group(1)}-{int(dm.group(2)):02d}-01"
        cert = line[:500]
        org = ""
        if "산업인력" in line or "협회" in line or "공단" in line:
            parts = line.split()
            if len(parts) >= 2:
                org = parts[-1][:500]
        rows.append(
            {
                "certificate": cert,
                "organization": org or None,
                "issue_date_raw": issue_raw,
                "certificate_number": None,
            }
        )
    return rows[:30]




def _strip_intro_question_head_markers(title: str) -> str:
    """자소서 문항 제목 앞 불릿(■ 등) 및 여분 공백 제거."""
    s = _nfc((title or "").strip())
    if not s:
        return ""
    s = re.sub(r"^[■▪●□▢▣・‧•∙]+\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _unwrap_square_heading_label(title: str) -> str:
    """표시용 문항 제목에서 바깥 [ ] 한 겹만 제거 (예: [자유서술] → 자유서술)."""
    s = _nfc((title or "").strip())
    if len(s) < 3 or not (s.startswith("[") and s.endswith("]")):
        return s
    inner = _nfc(s[1:-1]).strip()
    return inner if inner else s


def _sanitize_intro_question_title(raw: str, max_chars: int = 500) -> str:
    """문항 문자열 표시 정리: [ ] 제거, 선두 번호(1.) 제거, ■ 불릿 제거, 공백 정리."""
    t = _nfc((raw or "").strip())
    if not t:
        return ""
    t = re.sub(r"\s+", " ", t).strip()
    t = _unwrap_square_heading_label(t)
    t = re.sub(r"^[0-9]+\s*[.)．、]\s*", "", t)
    t = _strip_intro_question_head_markers(t)
    if len(t) > max_chars:
        t = t[:max_chars].strip()
    return t


def _intro_question_match_key(title: str) -> str:
    """질문 제목 비교용: 공백 제거·불릿 제거 후 NFC."""
    return re.sub(r"\s+", "", _strip_intro_question_head_markers(_nfc(title or "")))


def _strip_intro_answer_duplicate_question_line(
    answer: str, qh_clean: str | None
) -> str:
    """
    답변 맨 앞 줄이 질문 필드와 동일한 제목(■·번호 등 포함)이면 한 줄 제거.
    표 셀에서 헤더와 본문이 한 덩어리로 잡힐 때 question·answer 중복을 막는다.
    """
    if not answer or not qh_clean:
        return answer
    q_key = _intro_question_match_key(qh_clean)
    if not q_key:
        return answer
    t = _nfc(answer.replace("\r", ""))
    lines = t.split("\n")
    i = 0
    while i < len(lines) and not lines[i].strip():
        i += 1
    if i >= len(lines):
        return answer
    first = lines[i].strip()
    first_key = _intro_question_match_key(first)
    if first_key and first_key == q_key:
        rest = "\n".join(lines[i + 1 :]).lstrip("\n")
        return rest
    return answer


def _strip_trailing_orphan_heading_bullets(answer: str) -> str:
    """답변 끝의 '■'만 있는 줄 등 제목 잔여 불릿 제거."""
    if not answer:
        return answer
    lines = _nfc(answer.replace("\r", "")).split("\n")
    while lines:
        s = lines[-1].strip()
        if not s:
            lines.pop()
            continue
        if re.fullmatch(r"[■▪●・‧•∙□▢▣]+", s):
            lines.pop()
            continue
        break
    return "\n".join(lines).rstrip()


def _infer_question_freeform_intro_blob(norm_source: str, block: str) -> str | None:
    """[자유서술]로 잘린 통째 블록이면 문항 이름을 채운다 ([]는 포함하지 않음)."""
    nt = norm_source.replace("\r", "").strip()
    blob = block.strip()
    if len(blob) < 10:
        return None

    for m in re.finditer(r"\[\s*자유\s*서술\s*\]", nt):
        after = nt[m.end() :].lstrip()
        if not after:
            continue
        cb = re.sub(r"\s+", "", _nfc(blob))
        ca = re.sub(r"\s+", "", _nfc(after))
        take = min(len(cb), len(ca), 240)
        if take >= 40 and cb[:take] == ca[:take]:
            return "자유서술"

    return None


def _parse_vertical_pdf_resume_sections_strict(block: str) -> list[dict] | None:
    b = (block or "").strip()
    if len(b) < 30:
        return None
    if not re.search(r"(?ms)^\s*성\s+장\s*\n\s*과\s+정\s*", b):
        return None
    if not re.search(r"지\s*원\s*동\s*기", b) or not re.search(r"입\s*사\s*후", b):
        return None

    out: list[dict] = []
    m1 = re.search(
        r"(?ms)^\s*성\s+장\s*\n\s*과\s+정\s*\n(.+?)(?=^\s*성\s+격\s*$)",
        b,
    )
    if not m1:
        return None
    out.append({"question": "성장과정", "answer": m1.group(1).strip()[:20000]})

    m2 = re.search(
        r"(?ms)^\s*성\s+격\s*\n\s*및\s*\n\s*특기사항\s*\n(.+?)(?=지\s*원\s*동\s*기)",
        b,
        re.DOTALL,
    )
    if not m2:
        return None
    out.append({"question": "성격 및 특기사항", "answer": m2.group(1).strip()[:20000]})

    m3 = re.search(
        r"(?ms)지\s*원\s*동\s*기(.+?)입\s*사\s*후",
        b,
        re.DOTALL,
    )
    if not m3:
        return None
    out.append({"question": "지원동기", "answer": m3.group(1).strip()[:20000]})

    m4 = re.search(
        r"(?ms)입\s*사\s*후[^\n]{0,6}\s*\n\s*포부\s*\n(.+)",
        b,
        re.DOTALL,
    )
    if not m4:
        m4 = re.search(
            r"(?ms)입\s*사\s*후[^\n]+\n[^\n]{0,6}포부\s*\n(.+)",
            b,
            re.DOTALL,
        )
    if not m4:
        return None
    out.append({"question": "입사 후 포부", "answer": m4.group(1).strip()[:20000]})
    return out if len(out) >= 2 else None


def _parse_vertical_pdf_resume_sections_loose(block: str) -> list[dict] | None:
    """셀 순서 깨진 PDF: '지원동기' 위치까지를 앞머리로 보고 성장/성격 블록을 나눔."""
    b = (block or "").strip()
    if len(b) < 30:
        return None
    kj = re.search(r"(?ms)지\s*원\s*동\s*기", b)
    if not kj:
        return None
    ip_rest = re.search(r"입\s*사\s*후", b[kj.end() :])
    if not ip_rest:
        return None
    ip_abs = kj.end() + ip_rest.start()
    kj_body = b[kj.end() : ip_abs].strip()[:20000]
    head = b[: kj.start()].strip()
    sj_m = re.search(r"(?m)^\s*성\s+격\s*$", head)
    if sj_m:
        grow = head[: sj_m.start()].strip()[:20000]
        sj_block = head[sj_m.start() :].strip()[:20000]
        grow = re.sub(r"(?ms)^\s*성\s+장\s*$|^\s*과\s+정\s*$", "", grow).strip()
        sj_block = re.sub(
            r"(?ms)^\s*성\s+격\s*$|^\s*및\s*$|^\s*특기사항\s*$", "", sj_block
        ).strip()
    else:
        grow = head[:20000]
        sj_block = ""

    tail_block = b[ip_abs:].strip()
    lines = tail_block.split("\n")
    i = 0
    while i < len(lines) and not lines[i].strip():
        i += 1
    if i < len(lines) and re.match(r"(?i)^\s*입\s*사\s*후", lines[i].strip()):
        i += 1
    fourth_lines = lines[i:]
    while fourth_lines and not fourth_lines[0].strip():
        fourth_lines = fourth_lines[1:]
    if fourth_lines and re.match(r"(?i)^\s*포부\s*$", fourth_lines[-1].strip()):
        fourth_lines = fourth_lines[:-1]
    tail_answer = "\n".join(fourth_lines).strip()[:20000]

    rows: list[dict] = []
    if grow:
        rows.append({"question": "성장과정", "answer": grow})
    if sj_block:
        rows.append({"question": "성격 및 특기사항", "answer": sj_block})
    if kj_body:
        rows.append({"question": "지원동기", "answer": kj_body})
    if tail_answer:
        rows.append({"question": "입사 후 포부", "answer": tail_answer})
    return rows if len(rows) >= 2 else None




_INTRO_HDR_LINE_FULL = re.compile(
    r"(?m)^\s*((?:\d+[.)]\s+[^\n]{1,260})|(?:[■▪●□▢▣]+\s*\S[^\n]{1,420}))\s*$"
)


def _sanitize_vertical_intro_answer_rows(rows: list[dict]) -> list[dict]:
    """PDF 세로 OCR 보정 경로 결과에 동일 헤더 잔복·끝 ■ 줄 제거."""
    out = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        qh_raw = r.get("question")
        qh_clean = _nfc(qh_raw).strip() if qh_raw is not None else None
        if qh_clean == "":
            qh_clean = None
        ans = r.get("answer") or ""
        ans = _strip_intro_answer_duplicate_question_line(ans, qh_clean)
        ans = _strip_trailing_orphan_heading_bullets(ans)
        if ans:
            out.append({"question": qh_clean, "answer": ans[:20000]})
    return out


def _parse_intro_blocks(block: str, normalized_source: str | None = None) -> list[dict]:
    """번호(1.)·■항목 줄마다 문항 헤더 + 본문. 구분 불가하면 한 줄로 저장."""
    b = (block or "").strip()
    if len(b) < 10:
        return []
    # 1) PDF 한 글자씩 줄바꿈된 고정 순서 블록 (■ 문항 줄이 없어도 잡히는 형식)
    strict_vertical = _parse_vertical_pdf_resume_sections_strict(b)
    if strict_vertical:
        return strict_vertical

    headings = []
    for m in _INTRO_HDR_LINE_FULL.finditer(b):
        headings.append((m.start(), m.end(), m.group(1).strip()))
    headings.sort(key=lambda x: x[0])
    rows_out: list[dict] = []

    # 2) 표준 ■·번호 문항 줄 (DOCX·일반 PDF). 세로 OCR loose보다 우선 — loose가 지원동기 문자만 보고 한 행으로 뭉개는 문제 방지
    if len(headings) >= 2:
        for idx, (_, end_pos, qh) in enumerate(headings):
            nxt_start = headings[idx + 1][0] if idx + 1 < len(headings) else len(b)
            ans = _nfc(b[end_pos:nxt_start]).strip()
            qh_clean = _sanitize_intro_question_title(qh)
            ans = _strip_intro_answer_duplicate_question_line(ans, qh_clean)
            ans = _strip_trailing_orphan_heading_bullets(ans)
            if ans:
                rows_out.append(
                    {"question": qh_clean if qh_clean else None, "answer": ans[:20000]}
                )
            if len(rows_out) >= 30:
                break
        if rows_out:
            return rows_out

    if len(headings) == 1:
        _, end_pos, qh = headings[0]
        qh_clean = _sanitize_intro_question_title(qh)
        ans = _nfc(b[end_pos:].strip())[:20000]
        ans = _strip_intro_answer_duplicate_question_line(ans, qh_clean)
        ans = _strip_trailing_orphan_heading_bullets(ans)
        return [{"question": qh_clean if qh_clean else None, "answer": ans}]

    # 3) ■ 줄이 없거나 OCR로 줄이 깨진 PDF
    loose_vertical = _parse_vertical_pdf_resume_sections_loose(b)
    if loose_vertical:
        return _sanitize_vertical_intro_answer_rows(loose_vertical)

    q_blob = (
        _infer_question_freeform_intro_blob(normalized_source, b)
        if normalized_source
        else None
    )
    return [{"question": q_blob, "answer": b[:20000]}]


def _clip_profile_scalar(s: str, max_chars: int) -> str:
    """표 한 칸·라벨 뒤 짧은 값만. 다른 열(|, 취미/특기) 또는 긴 본문은 잘라냄."""
    v = _nfc((s or "").strip())
    if not v:
        return ""
    v = re.sub(r"\s+", " ", v)
    for sep in ("|", "｜"):
        if sep in v:
            v = v.split(sep)[0].strip()
    if re.search(r"취미|특기", v):
        v = re.split(r"(?:취미|특기)", v, maxsplit=1)[0].strip().rstrip("/").strip()
    if len(v) > max_chars:
        v = v[:max_chars].strip()
    return v.strip()


_SCALAR_HEADER_JUNK = frozenset({"대상여부", "1.대상2.비대상", "해당여부"})


def _scalar_value_clean(key: str, raw: str) -> str:
    """DB에 들어가기엔 과한 문단·헤더 잡음이면 버림."""
    s = raw.strip()
    if not s:
        return ""
    collapsed = re.sub(r"\s+", "", s)
    if collapsed in _SCALAR_HEADER_JUNK:
        return ""
    # 희망근무지에 보훈·장애·취미 행이 합쳐진 경우
    if key == K_DESIRED_LOCATION:
        if len(s) > 52 and re.search(r"(보훈|장애\s*인|미해당\s*취미|취미\s*/\s*특기)", s):
            return ""
    if len(s) > 120:
        bad = ("저는", "입사 후", "지원하였", "예상치", "특히", "삼성전자는", "서비스", "복잡한")
        if sum(1 for p in bad if p in s) >= 2 or (len(s) > 280 and "," in s and "입니다" in s):
            return ""
    if len(s.split()) > 35 and "," in s and "저" in s[:20]:
        return ""
    if key == K_VETERAN_ELIGIBILITY and re.fullmatch(r"[\s\d.,·\[\]①②③]+$", collapsed):
        return ""
    return _clip_profile_scalar(s, 120)


def _extract_disability_value(t: str) -> str:
    """
    장애: 인적표 '보훈/장애' 칸만 우선. 본문 '시스템 장애' 등은 제외.
    """
    m = re.search(
        r"보훈\s*[/／]\s*장애[^\n]{0,20}?\s*[：:|｜]?\s*([^\n|｜]{1,64})",
        t,
        re.MULTILINE,
    )
    if m:
        return _scalar_value_clean(K_DISABILITY, m.group(1))
    m = re.search(
        r"(?<!시스템 )(?<!네트워크 )"
        r"장애(?:\s*인(?:\([^)]*\))?)?\s*[:：]\s*([^\n\|｜]{1,64})",
        t,
    )
    if m:
        raw = _clip_profile_scalar(m.group(1), 80)
        if len(raw) > 90:
            return ""
        return _scalar_value_clean(K_DISABILITY, raw)
    return ""


def _extract_veteran_value(t: str) -> str:
    """보훈: 표·라벨 뒤 짧은 분류만 (표 헤더 '대상여부' 단독 추출 제외)."""
    for pat in (
        r"(?:국가\s*)?보훈(?:\s*여부)?[^\n:：]{0,12}[：:｜│]\s*([^｜|\n]{1,64})",
        r"국가\s*보훈\s*(?:대상)?[^\n:：]{0,12}[：:｜│]\s*([^｜|\n]{1,64})",
        r"보훈(?:\s*등록)?(?:\s*번호)?[^\n:：]{0,8}[：:｜│]\s*([^｜|\n]{1,64})",
    ):
        m = re.search(pat, t)
        if not m:
            continue
        v = _scalar_value_clean(K_VETERAN_ELIGIBILITY, m.group(1))
        if v and re.sub(r"\s+", "", v) not in _SCALAR_HEADER_JUNK:
            return v
    num = re.search(r"보훈\s*(?:등록|증)?\s*번호\s*[：:]\s*(\d[\d\s\-]{5,16})", t)
    if num:
        return _scalar_value_clean(K_VETERAN_ELIGIBILITY, num.group(1).strip())
    return ""


def _extract_desired_location_value(t: str) -> str:
    m = re.search(r"희망\s*근무지\s*(?:[^\n\:：]{0,12}?)?[:：]\s*([^\n]{1,100})", t)
    if m:
        return _scalar_value_clean(K_DESIRED_LOCATION, m.group(1))
    # 일진식: 표에서 '희망근무지 … | …' 줄
    m = re.search(
        r"희망근무[^\n]{0,12}?\s*[：:|｜]\s*([^｜|\n]{1,80})",
        t,
    )
    if m:
        return _scalar_value_clean(K_DESIRED_LOCATION, m.group(1))
    return ""


def _extract_supplement_scalars(text: str) -> dict[str, str]:
    out: dict[str, str] = {}
    t = text.replace("\r", "")
    gm = re.search(r"성별\s*[:：]?\s*(남성|여성|남|여|男|女)", t)
    if gm:
        g = gm.group(1)
        out[K_GENDER] = "여" if g in ("여", "女", "여성") else "남"

    dv = _extract_disability_value(t)
    if dv:
        out[K_DISABILITY] = dv

    vv = _extract_veteran_value(t)
    if vv:
        out[K_VETERAN_ELIGIBILITY] = vv

    wm = _extract_desired_location_value(t)
    if wm:
        out[K_DESIRED_LOCATION] = wm

    sm = re.search(r"희망\s*연봉\s*[:：]?\s*([^\n]{1,100})", t)
    if sm:
        out[K_DESIRED_SALARY] = _clip_profile_scalar(sm.group(1), 100)

    return out


def _merge_supplement_from_plaintext(text: str) -> dict:
    """병역·자격·자기소개 블록 + 기본 스칼라 (성별·보훈·장애·희망지/연봉)."""
    if not text or not text.strip():
        return {
            K_MILITARY_ROWS: [],
            K_QUALIFICATION_ROWS: [],
            K_STATEMENT_ROWS: [],
            **{k: "" for k in _SUPPLEMENT_SCALAR_KEYS},
        }
    common_end = (
        "학력사항",
        "경력사항",
        "자격사항",
        "병역사항",
        "자기소개서",
        "자기소개",
        "어학",
        "외국어",
        "해외경험",
        "수상실적",
        "수상",
        "OA",
        "포트폴리오",
    )
    intro_slice_ends = common_end + (
        "입사지원서",
        "[인적사항]",
        "상기 사항은",
    )
    mil_block = _slice_section_text(text, ("병역사항", "병 역"), common_end)
    qual_block = _slice_section_text(text, ("자격사항", "자 격"), common_end)
    nt_intro = _normalize_intro_header_spacing(text.replace("\r", ""))
    intro_groups = (
        ("자기소개서", "자기소개", "[자기소개서]"),
        ("[자유서술]", "자유서술"),
    )

    intro_block = _slice_intro_text_longest_candidate(nt_intro, intro_groups, intro_slice_ends)

    mil_rows = _parse_military_from_block(mil_block)
    qual_rows = _parse_qualification_from_block(qual_block)
    intro_rows = _parse_intro_blocks(intro_block, nt_intro)

    scalars = {k: "" for k in _SUPPLEMENT_SCALAR_KEYS}
    scalars.update(_extract_supplement_scalars(text))

    return {
        **scalars,
        K_MILITARY_ROWS: mil_rows,
        K_QUALIFICATION_ROWS: qual_rows,
        K_STATEMENT_ROWS: intro_rows,
    }


def _education_rows_from_docx_table(edu_rows: list[list[str]] | None) -> list[dict]:
    """DOCX 학력 표 → educations 테이블용 행 목록."""
    if not edu_rows or len(edu_rows) < 2:
        return []
    header = edu_rows[0]
    school_col = _find_col(header, ["학교명", "학교"])
    major_col = _find_col(header, ["전공", "학과명"])
    period_col = _find_col(header, ["재학기간", "기간"])
    level_col = _find_col(header, ["학력구분", "구분"])
    loc_col = _find_col(header, ["소재지"])
    grade_col = _find_col(header, ["학점", "평균학점"])
    out: list[dict] = []
    for row in edu_rows[1:]:
        row_cells = [_nfc(x) for x in row]
        row_text = " ".join(row_cells)
        if not any(kw in row_text for kw in ["대학교", "대학원", "대학", "고등학교"]):
            continue
        school = ""
        if school_col >= 0 and school_col < len(row_cells):
            school = row_cells[school_col]
        if not school or school in _EDU_SKIP:
            sm = _SCHOOL.search(row_text)
            school = sm.group(1).replace(" ", "") if sm else ""
        else:
            sm = _SCHOOL.search(school)
            if sm:
                school = sm.group(1).replace(" ", "")
        major = ""
        if major_col >= 0 and major_col < len(row_cells):
            major = row_cells[major_col]
            if major in _EDU_SKIP:
                major = ""
        period = ""
        if period_col >= 0 and period_col < len(row_cells):
            period = row_cells[period_col]
            if period in _EDU_SKIP:
                period = ""
        if not period:
            dr = _DATE_RANGE.search(row_text)
            if dr:
                period = dr.group(1)
        if not period:
            yms = _YEARMONTH.findall(row_text)
            if len(yms) >= 2:
                period = f"{yms[0]}~{yms[-1]}"
            elif yms:
                period = yms[0]
        level = ""
        if level_col >= 0 and level_col < len(row_cells):
            level = row_cells[level_col]
        level_for_infer = level or None
        location = ""
        if loc_col >= 0 and loc_col < len(row_cells):
            location = row_cells[loc_col]
        grade = ""
        if grade_col >= 0 and grade_col < len(row_cells):
            grade = row_cells[grade_col]
        a0, a1 = parse_education_period_raw_pair(period, row_text)
        comp = infer_education_level(school, major, level_for_infer, row_text)
        out.append(
            {
                "school_name": school or None,
                "department": major or None,
                "completion_status": comp,
                "period_raw": period or None,
                "location": location or None,
                "grade": grade or None,
                "attendance_start_period_raw": a0,
                "attendance_end_period_raw": a1,
            }
        )
    return [_refine_education_row_dict(r) for r in out]


def _qualification_rows_from_docx_table_headerless(
    qual_rows: list[list[str]],
) -> list[dict]:
    """첫 열 YYYY.MM·둘째 열 자격명·(셋째 발행처) — 헤더 행 없음."""
    out: list[dict] = []
    for row in qual_rows:
        row_cells = [_nfc(x) for x in row]
        if len(row_cells) < 2:
            continue
        c0 = row_cells[0].strip()
        if not _cell_is_yyyy_mm_only(c0):
            continue
        cert = row_cells[1].strip()
        if not cert or _qual_line_is_noise(cert):
            continue
        m = _CELL_YYYY_MM_ONLY.match(c0)
        if not m:
            continue
        y, mo = m.group(1), int(m.group(2))
        issue_raw = f"{y}-{mo:02d}-01"
        org = ""
        if len(row_cells) > 2:
            org = row_cells[2].strip()
        out.append(
            {
                "certificate": cert[:500],
                "organization": org or None,
                "issue_date_raw": issue_raw,
                "certificate_number": None,
            }
        )
        if len(out) >= 50:
            break
    return out


def _qualification_rows_from_docx_table(
    qual_rows: list[list[str]] | None,
) -> list[dict]:
    """DOCX 자격(면허) 표 → qualifications 테이블용. 영문 Certificate / Certification 헤더 대소문자 무관."""
    if not qual_rows:
        return []
    header = [_nfc(c) for c in qual_rows[0]]
    cert_col = _find_col(
        header,
        [
            "certification",
            "certificate name",
            "certificate",
            "qualification name",
            "qualification",
            "credential",
            "license name",
            "license",
            "자격증명",
            "자격증",
            "자격명",
            "자격종류",
            "합격종목",
            "종목",
        ],
    )
    if cert_col < 0 and _is_headerless_qual_grid(qual_rows):
        return _qualification_rows_from_docx_table_headerless(qual_rows)
    if len(qual_rows) < 2:
        return []
    if cert_col < 0 and len(header) >= 2:
        h0 = header[0].strip().lower()
        if any(x in h0 for x in ("no", "no.", "#", "연번", "순번", "번호", "□")) or len(
            header[0].strip()
        ) <= 2:
            cert_col = 1
    if cert_col < 0:
        cert_col = 0

    org_col = _find_col(
        header,
        ["발행처", "발급처", "발급기관", "기관", "organization", "issuer", "issued by"],
    )
    date_col = _find_col(
        header,
        [
            "취득일",
            "합격일",
            "발급일",
            "issue date",
            "date acquired",
            "acquired",
            "취득",
        ],
    )
    num_col = _find_col(
        header,
        [
            "자격증번호",
            "증서번호",
            "certificate number",
            "license number",
            "등록번호",
            "관리번호",
        ],
    )

    out: list[dict] = []
    for row in qual_rows[1:]:
        row_cells = [_nfc(x) for x in row]
        if cert_col >= len(row_cells):
            continue
        cert = (row_cells[cert_col] or "").strip()
        if not cert or _qual_line_is_noise(cert):
            continue
        org = ""
        if org_col >= 0 and org_col < len(row_cells):
            org = (row_cells[org_col] or "").strip()
        date_cell = ""
        if date_col >= 0 and date_col < len(row_cells):
            date_cell = (row_cells[date_col] or "").strip()
        issue_raw = None
        dm = re.search(r"(20\d{2})[.\-/년]\s*(\d{1,2})", date_cell)
        if dm:
            issue_raw = f"{dm.group(1)}-{int(dm.group(2)):02d}-01"
        cert_num = None
        if num_col >= 0 and num_col < len(row_cells):
            cert_num = (row_cells[num_col] or "").strip() or None
        out.append(
            {
                "certificate": cert[:500],
                "organization": org or None,
                "issue_date_raw": issue_raw,
                "certificate_number": cert_num,
            }
        )
        if len(out) >= 50:
            break
    return out


def _experience_rows_from_career_entries(entries: list[dict]) -> list[dict]:
    rows = []
    for e in entries:
        period = (e.get("period") or "").strip()
        s0, s1 = _parse_period_to_raw_pair(period)
        rows.append(
            {
                "company": _empty_strip(e.get("company")),
                "role": _empty_strip(e.get("job")),
                "job_title": None,
                "salary_raw": None,
                "reason_for_leaving": None,
                "employment_start_period_raw": s0,
                "employment_end_period_raw": s1,
            }
        )
    return rows


def _experience_rows_from_merged_lines(
    career_period: str,
    career_company: str,
    career_role: str,
) -> list[dict]:
    pp = [x.strip() for x in (career_period or "").split("\n") if x.strip()]
    cc = [x.strip() for x in (career_company or "").split("\n") if x.strip()]
    rr = [x.strip() for x in (career_role or "").split("\n") if x.strip()]
    n = max(len(pp), len(cc), len(rr))
    if n == 0:
        return []
    rows: list[dict] = []
    for i in range(n):
        period = pp[i] if i < len(pp) else (pp[-1] if pp else "")
        company = cc[i] if i < len(cc) else ""
        role = rr[i] if i < len(rr) else ""
        s0, s1 = _parse_period_to_raw_pair(period)
        rows.append(
            {
                "company": company or None,
                "role": role or None,
                "job_title": None,
                "salary_raw": None,
                "reason_for_leaving": None,
                "employment_start_period_raw": s0,
                "employment_end_period_raw": s1,
            }
        )
    return rows


def _finalize_parsed_record(parsed: dict) -> dict:
    """추출 실패/누락 시 안전값으로 보정"""
    required_keys = [
        K_NAME, K_BIRTH_DATE, K_CONTACT, K_EMAIL, K_ADDR,
        K_ORIGINAL_JOB_ROLE, K_FINAL_EDU, K_CAREER_PD, K_CAREER_CO, K_CAREER_ROLE,
    ]

    normalized: dict = {}
    for k in required_keys:
        v = parsed.get(k, "")
        if v is None:
            v = ""
        normalized[k] = str(v).strip()

    # 카테고리성 정보는 누락 시 '기타'로 보정
    if not normalized[K_ORIGINAL_JOB_ROLE]:
        normalized[K_ORIGINAL_JOB_ROLE] = "기타"
    if not normalized[K_FINAL_EDU]:
        normalized[K_FINAL_EDU] = "기타"
    if not normalized[K_CAREER_ROLE] and normalized[K_CAREER_PD]:
        normalized[K_CAREER_ROLE] = "기타"
    if not normalized[K_CAREER_CO] and normalized[K_CAREER_PD]:
        normalized[K_CAREER_CO] = "기타"

    for k in _SUPPLEMENT_SCALAR_KEYS:
        v = parsed.get(k, "")
        if v is None:
            normalized[k] = ""
        else:
            normalized[k] = str(v).strip()

    for k in _STRUCTURE_RESULT_KEYS:
        v = parsed.get(k)
        if isinstance(v, list):
            normalized[k] = v
        else:
            normalized[k] = []

    # 부가 필드는 기존 값 유지(extractionConfidence는 사용하지 않음)
    # 구조화/스칼라 보조 키는 위에서 처리 — 리스트(dict) 값은 문자열화하지 않음
    _skip = set(required_keys) | _SUPPLEMENT_SCALAR_KEYS | _STRUCTURE_RESULT_KEYS
    for k, v in parsed.items():
        if k in _skip:
            continue
        if k == K_EXTRACT_CONF:
            continue
        if isinstance(v, (list, dict)):
            normalized[k] = v
        else:
            normalized[k] = "" if v is None else str(v).strip()

    return normalized


def _extract_json_object(text: str) -> dict:
    """모델 응답에서 첫 JSON 객체를 추출"""
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            return json.loads(text[start:end + 1])
        except Exception:
            return {}
    return {}


def _should_anonymize_ai() -> bool:
    return os.getenv("AI_ANONYMIZE", "true").strip().lower() in {"1", "true", "yes", "on"}


def _mask_sensitive_text(text: str) -> tuple[str, dict[str, str]]:
    """AI 전송 전 민감정보를 토큰으로 치환"""
    masked = text
    token_map: dict[str, str] = {}
    seq = 1

    def _add_token(raw: str, kind: str) -> str:
        nonlocal seq
        raw = raw.strip()
        if not raw:
            return raw
        for tok, original in token_map.items():
            if original == raw:
                return tok
        token = f"__{kind}_{seq}__"
        seq += 1
        token_map[token] = raw
        return token

    # 1) 이메일
    email_pat = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
    for m in sorted(set(email_pat.findall(masked)), key=len, reverse=True):
        masked = masked.replace(m, _add_token(m, "EMAIL"))

    # 2) 전화번호
    phone_pat = re.compile(r"(?:010|011|016|017|018|019|02|0[3-6]\d|070)[-\s]?\d{3,4}[-\s]?\d{4}")
    for m in sorted(set(phone_pat.findall(masked)), key=len, reverse=True):
        masked = masked.replace(m, _add_token(m, "PHONE"))

    # 3) 생년월일/주민번호 유사 패턴
    birth_pat = re.compile(r"\b\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}\b")
    for m in sorted(set(birth_pat.findall(masked)), key=len, reverse=True):
        masked = masked.replace(m, _add_token(m, "BIRTH"))

    rrn_pat = re.compile(r"\b\d{6}[-\s]?\d{7}\b")
    for m in sorted(set(rrn_pat.findall(masked)), key=len, reverse=True):
        masked = masked.replace(m, _add_token(m, "PID"))

    # 4) 기본정보 기반 이름/주소
    basic = _extract_basic(masked)
    if basic.get(K_NAME):
        name = basic[K_NAME].strip()
        if len(name) >= 2:
            masked = masked.replace(name, _add_token(name, "NAME"))
    if basic.get(K_BIRTH_DATE):
        birth = basic[K_BIRTH_DATE].strip()
        if birth:
            masked = masked.replace(birth, _add_token(birth, "BIRTH"))
    if basic.get(K_ADDR):
        addr = basic[K_ADDR].strip()
        if len(addr) >= 4:
            masked = masked.replace(addr, _add_token(addr, "ADDR"))

    return masked, token_map


def _restore_tokens(value: str, token_map: dict[str, str]) -> str:
    restored = value
    for token, original in token_map.items():
        restored = restored.replace(token, original)
    return restored


def _call_openai(prompt: str) -> str:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY가 없습니다.")
    model = os.getenv("OPENAI_MODEL", "").strip()
    if not model:
        raise RuntimeError("OPENAI_MODEL이 없습니다.")
    payload = {
        "model": model,
        "temperature": 0,
        "messages": [
            {"role": "system", "content": "당신은 이력서 정보를 JSON으로 추출하는 엔진입니다."},
            {"role": "user", "content": prompt},
        ],
    }
    req = urllib.request.Request(
        url="https://api.openai.com/v1/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8", errors="replace"))
    return data["choices"][0]["message"]["content"]


def _call_anthropic(prompt: str) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY가 없습니다.")
    model = os.getenv("ANTHROPIC_MODEL", "").strip()
    if not model:
        raise RuntimeError("ANTHROPIC_MODEL이 없습니다.")
    payload = {
        "model": model,
        "max_tokens": 1200,
        "temperature": 0,
        "messages": [{"role": "user", "content": prompt}],
    }
    req = urllib.request.Request(
        url="https://api.anthropic.com/v1/messages",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8", errors="replace"))
    blocks = data.get("content", [])
    text_blocks = [b.get("text", "") for b in blocks if b.get("type") == "text"]
    return "\n".join(text_blocks).strip()


def _call_gemini(prompt: str) -> str:
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY가 없습니다.")
    model = os.getenv("GEMINI_MODEL", "").strip()
    if not model:
        raise RuntimeError("GEMINI_MODEL이 없습니다.")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    payload = {
        "generationConfig": {"temperature": 0},
        "contents": [{"parts": [{"text": prompt}]}],
    }
    req = urllib.request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8", errors="replace"))
    cands = data.get("candidates", [])
    if not cands:
        return ""
    parts = cands[0].get("content", {}).get("parts", [])
    return "\n".join(p.get("text", "") for p in parts if p.get("text")).strip()


def _should_use_ai() -> bool:
    return os.getenv("USE_AI", "false").strip().lower() in {"1", "true", "yes", "on"}


def _resolve_ai_providers() -> list[str]:
    raw = os.getenv("AI_PROVIDER", "").strip().lower()
    available: list[str] = []
    if os.getenv("OPENAI_API_KEY", "").strip():
        available.append("openai")
    if os.getenv("ANTHROPIC_API_KEY", "").strip():
        available.append("anthropic")
    if os.getenv("GEMINI_API_KEY", "").strip():
        available.append("gemini")

    if not raw:
        return available[:1]
    if raw in {"all", "ensemble", "multi"}:
        return available

    requested = [x.strip() for x in raw.split(",") if x.strip()]
    resolved = [p for p in requested if p in {"openai", "anthropic", "gemini"} and p in available]
    return resolved


def _aggregate_ai_results(results: list[tuple[str, dict]]) -> dict:
    if not results:
        return {}
    fields = [
        K_NAME, K_BIRTH_DATE, K_CONTACT, K_EMAIL, K_ADDR,
        K_ORIGINAL_JOB_ROLE, K_FINAL_EDU, K_CAREER_PD, K_CAREER_CO, K_CAREER_ROLE,
    ]
    merged: dict[str, str] = {}
    for field in fields:
        votes: dict[str, int] = {}
        for _, r in results:
            v = str(r.get(field, "")).strip()
            if not v:
                continue
            votes[v] = votes.get(v, 0) + 1
        if not votes:
            merged[field] = ""
            continue
        # 다수결 우선, 동률이면 길이가 긴 값 우선
        best = sorted(votes.items(), key=lambda kv: (kv[1], len(kv[0])), reverse=True)[0][0]
        merged[field] = best

    # 최종학력은 유효성 체크를 통과한 값만 유지
    if merged.get(K_FINAL_EDU) and not _is_valid_final_edu(merged[K_FINAL_EDU]):
        merged[K_FINAL_EDU] = ""
    return merged


def _ai_extract_fields(text: str, filename: str, filetype: str) -> dict:
    """AI API로 이력서 필드 추출"""
    providers = _resolve_ai_providers()
    if not providers:
        return {}

    prompt_text = text[:12000]
    token_map: dict[str, str] = {}
    if _should_anonymize_ai():
        prompt_text, token_map = _mask_sensitive_text(prompt_text)

    prompt = (
        "아래 텍스트에서 이력서 정보를 추출해 JSON만 반환하세요.\n"
        "규칙:\n"
        "1) 최종학력은 반드시 학교명+기간이 같이 있는 항목만 인정.\n"
        "2) '단과대학', '학교명' 같은 헤더 문구는 값으로 쓰지 말 것.\n"
        "3) 없는 값은 빈 문자열.\n"
        "4) 키는 정확히 아래만 사용.\n"
        "5) __NAME_n__, __PHONE_n__, __EMAIL_n__, __ADDR_n__, __BIRTH_n__, __PID_n__ 같은 토큰이 보이면 그대로 유지.\n\n"
        "{\n"
        f'  "{K_NAME}": "", "{K_BIRTH_DATE}": "", "{K_CONTACT}": "", "{K_EMAIL}": "", "{K_ADDR}": "",\n'
        f'  "{K_ORIGINAL_JOB_ROLE}": "", "{K_FINAL_EDU}": "", "{K_CAREER_PD}": "", "{K_CAREER_CO}": "", "{K_CAREER_ROLE}": ""\n'
        "}\n\n"
        f"[파일명] {filename}\n"
        f"[파일형식] {filetype}\n"
        "[문서텍스트]\n"
        f"{prompt_text}"
    )

    parsed_results: list[tuple[str, dict]] = []
    for provider in providers:
        try:
            if provider == "openai":
                raw = _call_openai(prompt)
            elif provider == "anthropic":
                raw = _call_anthropic(prompt)
            else:
                raw = _call_gemini(prompt)
            parsed = _extract_json_object(raw)
            if isinstance(parsed, dict):
                parsed_results.append((provider, parsed))
        except Exception as e:
            print(f"[WARN] AI 추출 실패({provider}): {e}")

    if not parsed_results:
        return {}
    allowed = {
        K_NAME, K_BIRTH_DATE, K_CONTACT, K_EMAIL, K_ADDR,
        K_ORIGINAL_JOB_ROLE, K_FINAL_EDU, K_CAREER_PD, K_CAREER_CO, K_CAREER_ROLE,
    }
    normalized = []
    for provider, parsed in parsed_results:
        normalized.append((provider, {k: str(parsed.get(k, "")).strip() for k in allowed}))

    result = _aggregate_ai_results(normalized) if len(normalized) > 1 else normalized[0][1]
    if token_map:
        for k, v in list(result.items()):
            result[k] = _restore_tokens(v, token_map)
    return result


def _is_valid_final_edu(value: str) -> bool:
    if _is_bad_edu_value(value):
        return False
    return bool(re.search(r"\(\d{4}[.\-]\d{1,2}\s*[~\-]\s*\d{4}[.\-]\d{1,2}", value))


def _normalize_for_match(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def _exists_in_source(candidate: str, source_text: str) -> bool:
    if not candidate.strip():
        return False
    c = _normalize_for_match(candidate)
    s = _normalize_for_match(source_text)
    if c in s:
        return True
    # 기간 표기 차이(하이픈/~/점) 완화
    c2 = c.replace("~", "").replace("-", "").replace(".", "")
    s2 = s.replace("~", "").replace("-", "").replace(".", "")
    return c2 in s2


def _merge_ai_result(base: dict, ai: dict, source_text: str) -> dict:
    if not ai:
        return base
    merged = dict(base)
    for k in [K_NAME, K_BIRTH_DATE, K_CONTACT, K_EMAIL, K_ADDR, K_ORIGINAL_JOB_ROLE]:
        if not merged.get(k) and ai.get(k) and _exists_in_source(ai.get(k, ""), source_text):
            merged[k] = ai[k]
    if (
        ai.get(K_FINAL_EDU)
        and _exists_in_source(ai.get(K_FINAL_EDU, ""), source_text)
        and (_is_bad_edu_value(merged.get(K_FINAL_EDU, "")) and _is_valid_final_edu(ai[K_FINAL_EDU]))
    ):
        merged[K_FINAL_EDU] = ai[K_FINAL_EDU]
    for k in [K_CAREER_PD, K_CAREER_CO, K_CAREER_ROLE]:
        if not merged.get(k) and ai.get(k) and _exists_in_source(ai.get(k, ""), source_text):
            merged[k] = ai[k]
    merged.pop(K_EXTRACT_CONF, None)
    return merged


# ─── DOCX 직접 파싱 ─────────────────────────────────────────────────────────

def extract_from_docx(filepath: str, original_filename: str) -> dict:
    """python-docx로 DOCX 테이블 구조를 직접 파싱 (LibreOffice 불필요)"""
    try:
        from docx import Document as _Doc
    except ImportError:
        subprocess.run(["pip", "install", "python-docx", "-q"], check=False)
        from docx import Document as _Doc

    doc = _Doc(filepath)

    # ── 1. 모든 단락 & 테이블을 텍스트로 ──
    para_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    all_tables: list[list[list[str]]] = []
    for table in doc.tables:
        rows = []
        for row in table.rows:
            # 연속 병합 셀 중복 제거 후 저장
            raw = [c.text.strip().replace("\n", " ").replace("\r", " ") for c in row.cells]
            deduped: list[str] = []
            prev = None
            for cell in raw:
                if cell != prev:
                    deduped.append(cell)
                    prev = cell
            rows.append(deduped)
        all_tables.append(rows)

    # 전체 텍스트 (기본정보 파싱용) - 테이블을 먼저 배치해 구조화된 정보 우선 매칭
    table_flat = "\n".join(
        " ".join(cell for cell in row) for rows in all_tables for row in rows
    )
    full_text = table_flat + "\n" + para_text

    # ── 2. 기본정보: 셀 기반 먼저 추출 → 누락 필드만 텍스트 정규식으로 보완 ──
    # (셀 기반을 먼저 실행해야 병합 셀·사진 칸 오염을 차단할 수 있음)
    info: dict[str, str] = {
        K_NAME: "",
        K_BIRTH_DATE: "",
        K_CONTACT: "",
        K_EMAIL: "",
        K_ADDR: "",
        K_ORIGINAL_JOB_ROLE: "",
    }

    # 테이블 셀에서 기본 필드 직접 추출 (병합 셀 / 사진 칸 오염 방지)
    _PHOTO_SKIP = {"사진", "사  진", "사 진", "photo", "Photo"}

    def _is_noise_cell(v: str) -> bool:
        if not v.strip():
            return True
        if any(p in v for p in _PHOTO_SKIP):
            return True
        return _norm(v).lower() in (_NAME_NORM | _PHONE_NORM | _ADDR_NORM | _EMAIL_NORM | _BIRTH_NORM)

    def _pick_value(v: str, want_pattern: str | None = None) -> str:
        if _is_noise_cell(v):
            return ""
        if want_pattern:
            m = re.search(want_pattern, v)
            return m.group(1) if m else ""
        return v.strip()

    def _find_value_near_key(
        rows: list[list[str]],
        row_idx: int,
        key_idx: int,
        want_pattern: str | None = None,
    ) -> str:
        """키 셀 주변에서 값을 찾음(오른쪽 우선, 아래/위 폴백)"""
        current = rows[row_idx]

        # 1) 같은 행 오른쪽
        for j in range(key_idx + 1, len(current)):
            picked = _pick_value(current[j], want_pattern)
            if picked:
                return picked

        # 2) 바로 아래/위 동일 컬럼 (양식에 따라 값이 세로 배치되는 경우)
        for offset in (1, -1):
            r = row_idx + offset
            if 0 <= r < len(rows) and key_idx < len(rows[r]):
                picked = _pick_value(rows[r][key_idx], want_pattern)
                if picked:
                    return picked

        # 3) 바로 아래/위 오른쪽 인접 컬럼
        for offset in (1, -1):
            r = row_idx + offset
            if 0 <= r < len(rows):
                for j in range(key_idx + 1, len(rows[r])):
                    picked = _pick_value(rows[r][j], want_pattern)
                    if picked:
                        return picked

        return ""

    def _norm(s: str) -> str:
        return re.sub(r"\s+", "", s)

    _NAME_NORM   = {"성명", "이름", "name"}
    _PHONE_NORM  = {"연락처", "휴대폰", "핸드폰", "mobile", "전화번호", "모바일",
                    "휴대전화", "핸드폰"}
    _ADDR_NORM   = {"주소", "거주지", "address"}
    _EMAIL_NORM  = {"email", "e-mail", "이메일"}
    _BIRTH_NORM  = {"생년월일", "생년", "birthdate"}

    for rows in all_tables:
        for row_idx, row in enumerate(rows):
            for i, cell in enumerate(row):
                nk = _norm(cell).lower()
                if nk in _NAME_NORM and not info[K_NAME]:
                    info[K_NAME] = _find_value_near_key(rows, row_idx, i, r"([가-힣]{2,5})")
                elif nk in _PHONE_NORM and not info[K_CONTACT]:
                    _mobile_pat   = r"(?:010|011|016|017|018|019)[-\s]?\d{3,4}[-\s]?\d{4}"
                    _landline_pat = (
                        r"(?:02|031|032|033|041|042|043|044|051|052|053"
                        r"|054|055|061|062|063|064|070)[-\s]?\d{3,4}[-\s]?\d{4}"
                    )
                    # 행 전체에서 휴대폰 번호 먼저 탐색, 없으면 지역번호
                    row_text = " ".join(row)
                    mob_m  = re.search(_mobile_pat,   row_text)
                    land_m = re.search(_landline_pat, row_text)
                    if mob_m:
                        info[K_CONTACT] = mob_m.group()
                    elif land_m:
                        info[K_CONTACT] = land_m.group()
                    else:
                        near_phone = _find_value_near_key(rows, row_idx, i)
                        mob_m = re.search(_mobile_pat, near_phone)
                        land_m = re.search(_landline_pat, near_phone)
                        if mob_m:
                            info[K_CONTACT] = mob_m.group()
                        elif land_m:
                            info[K_CONTACT] = land_m.group()
                elif nk in _ADDR_NORM and not info[K_ADDR]:
                    raw_addr = _find_value_near_key(rows, row_idx, i)
                    # 사진 칸 텍스트 후처리 제거: "사  진 (3cm x 4cm)" 등
                    raw_addr = re.sub(r'\s*사\s*진\s*[\(\（].*?[\)\）]', '', raw_addr).strip()
                    raw_addr = re.sub(r'\s*Photo\s*[\(\（].*?[\)\）]', '', raw_addr, flags=re.IGNORECASE).strip()
                    info[K_ADDR] = raw_addr
                elif nk in _EMAIL_NORM and not info[K_EMAIL]:
                    info[K_EMAIL] = _find_value_near_key(
                        rows, row_idx, i, r"([a-zA-Z0-9._%+\-가-힣]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})"
                    )
                elif nk in _BIRTH_NORM and not info[K_BIRTH_DATE]:
                    info[K_BIRTH_DATE] = _find_value_near_key(rows, row_idx, i, r"(\d{4}[.\-/]\d{2}[.\-/]\d{2})")

    # 셀 기반으로 못 찾은 필드만 텍스트 정규식으로 보완
    text_info = _extract_basic(full_text)
    for k, v in text_info.items():
        if not info.get(k):
            info[k] = v

    # ── 3. 학력/경력/자격 테이블 찾기 (헤더 대소문자 무관, certificate·certification 등 인식)
    EDU_HDR_KWS = ["학교명", "재학기간", "학력구분", "전공", "졸업구분", "학력", "기간", "학과명"]
    CAREER_HDR_KWS = ["회사명", "재직기간", "담당", "직무", "경력", "업무", "재직부서"]
    QUAL_HDR_KWS = [
        "자격증",
        "자격명",
        "자격종류",
        "자격사항",
        "취득일",
        "합격일",
        "발행처",
        "certificate",
        "certification",
        "credential",
        "qualification",
        "license",
        "등급",
        "종류",
    ]

    edu_rows: list[list[str]] | None = None
    career_rows: list[list[str]] | None = None
    qual_table_rows: list[list[str]] | None = None

    for rows in all_tables:
        if not rows:
            continue
        hdr = " ".join(rows[0])
        hdr_lc = hdr.lower()
        edu_score = sum(1 for kw in EDU_HDR_KWS if kw.lower() in hdr_lc)
        career_score = sum(1 for kw in CAREER_HDR_KWS if kw.lower() in hdr_lc)
        qual_score = sum(1 for kw in QUAL_HDR_KWS if kw.lower() in hdr_lc)
        if edu_score >= 2 and edu_rows is None:
            edu_rows = rows
        if career_score >= 2 and career_rows is None:
            career_rows = rows
        if qual_score >= 2 and qual_table_rows is None:
            qual_table_rows = rows

    # 헤더 없이 첫 열만 연·월(2026.01)인 자격·어학 표 (Certificate & Language 아래 등)
    if qual_table_rows is None:
        best_hl = 0
        best_tbl: list[list[str]] | None = None
        for rows in all_tables:
            if not rows or len(rows) < 2:
                continue
            hc = _headerless_qual_match_count(rows)
            if hc >= 2 and hc > best_hl:
                best_hl = hc
                best_tbl = rows
        if best_tbl is not None:
            qual_table_rows = best_tbl

    # ── 4. 학력 파싱 ──
    final_edu = ""
    edu_date = ""

    if edu_rows:
        header = edu_rows[0]
        best_score = -1

        school_col = _find_col(header, ["학교명", "학교"])
        major_col  = _find_col(header, ["전공", "학과명"])
        period_col = _find_col(header, ["재학기간", "기간"])
        status_col = _find_col(header, ["졸업구분"])   # '졸업' 단독 컬럼은 날짜값이므로 제외
        level_col  = _find_col(header, ["학력구분", "구분"])

        for row in edu_rows[1:]:
            row_text = " ".join(row)
            if not any(kw in row_text for kw in ["대학교", "대학원", "대학"]):
                continue

            # 학교명
            school = ""
            if school_col >= 0 and school_col < len(row):
                school = row[school_col].strip()
            if not school or school in _EDU_SKIP:
                sm = _SCHOOL.search(row_text)
                school = sm.group(1).replace(" ", "") if sm else ""
            else:
                sm = _SCHOOL.search(school)
                if sm:
                    school = sm.group(1).replace(" ", "")

            # 전공
            major = ""
            if major_col >= 0 and major_col < len(row):
                major = row[major_col].strip()
                if major in _EDU_SKIP:
                    major = ""

            # 기간: period 컬럼 → 날짜 범위 패턴 → 별도 입학/졸업 셀 조합
            period = ""
            if period_col >= 0 and period_col < len(row):
                period = row[period_col].strip()
                if period in _EDU_SKIP:
                    period = ""
            if not period:
                dr = _DATE_RANGE.search(row_text)
                if dr:
                    period = dr.group(1)
            if not period:
                # 일반사무 포맷: 입학 연월 + 졸업 연월이 별도 셀
                yms = _YEARMONTH.findall(row_text)
                if len(yms) >= 2:
                    period = f"{yms[0]}~{yms[-1]}"
                elif yms:
                    period = yms[0]
            candidate_period = period

            # 졸업 상태
            status = ""
            if status_col >= 0 and status_col < len(row):
                status = row[status_col].strip()
                # 날짜처럼 생긴 값(2016.02 등)이나 헤더값은 무시
                if status in _EDU_SKIP or status == "-" or _YEARMONTH.search(status):
                    status = ""
            if not status:
                status = next((gs for gs in _GRAD_STATUS if gs in row_text), "")

            score = _score_edu_candidate(school, major, status, candidate_period)
            if score > best_score:
                best_score = score
                edu_date = candidate_period
                final_edu = _build_edu_str(school, major, status, edu_date)

    # ── 5. 경력 파싱 (기간 + 회사 + 직무) ──
    _CAREER_HDR_VALS = {
        "회사명", "재직기간", "담당", "직무", "경력", "업무", "재직부서",
        "부서", "직위", "직급", "기간", "회사", "기관명", "구분", "활동기간",
    }
    career_entries: list[dict] = []

    if career_rows:
        header = career_rows[0]

        # 기간 컬럼
        period_col = _find_col(header, ["재직기간", "기간", "활동기간"])

        # 회사명 컬럼
        company_col = _find_col(header, ["회사명", "회사", "기관명"])

        # 직무 컬럼 (우선순위: 직무/담당 > 업무 > 부서/재직부서 > 직위/직급)
        job_col = _find_col(header, ["직무", "담당"])
        if job_col < 0:
            job_col = _find_col(header, ["업무"])
        if job_col < 0:
            job_col = _find_col(header, ["부서", "재직부서"])
        if job_col < 0:
            job_col = _find_col(header, ["직위", "직급"])

        for row in career_rows[1:]:
            row_text = " ".join(row)

            # 기간
            period = ""
            if period_col >= 0 and period_col < len(row):
                period = row[period_col].strip()
                if period in _EDU_SKIP or not re.search(r"\d{4}", period):
                    period = ""
            if not period:
                dr = _DATE_RANGE.search(row_text)
                if dr:
                    period = dr.group(1)
            if not period:
                continue  # 기간 없으면 경력 행 아님

            # 회사명
            company = ""
            if company_col >= 0 and company_col < len(row):
                company = row[company_col].strip()
                if company in _CAREER_HDR_VALS:
                    company = ""

            # 직무
            job = ""
            if job_col >= 0 and job_col < len(row):
                job = row[job_col].strip()
                if job in _CAREER_HDR_VALS:
                    job = ""
                if len(job) > 60:
                    job = job[:60] + "…"

            career_entries.append({"period": period, "company": company, "job": job})

    career_period  = "\n".join(e["period"]  for e in career_entries)
    career_company = "\n".join(e["company"] for e in career_entries)
    career_job     = "\n".join(e["job"]     for e in career_entries)

    # ── 6. 테이블에서 못 찾은 경우 텍스트 기반 폴백 ──
    all_dates = _DATE_RANGE.findall(full_text)
    clean_flat = re.sub(r"[ \t]+", " ", full_text)

    if _is_bad_edu_value(final_edu):
        final_edu, edu_date = _parse_edu_from_text(clean_flat, all_dates)

    if not career_period:
        career_period, career_company, career_job = _parse_career_from_text(
            clean_flat, all_dates, edu_date
        )

    # 주소 최종 정제: 사진 칸 텍스트 잔여물 제거 (어떤 경로로 설정됐든 무조건 적용)
    if info.get(K_ADDR):
        info[K_ADDR] = re.sub(r'\s*사\s*진.*$', '', info[K_ADDR]).strip()
        info[K_ADDR] = re.sub(r'\s*Photo.*$', '', info[K_ADDR], flags=re.IGNORECASE).strip()

    parsed = {
        **info,
        K_FINAL_EDU: final_edu,
        K_CAREER_PD: career_period,
        K_CAREER_CO: career_company,
        K_CAREER_ROLE: career_job,
        K_FILE_TYPE: os.path.splitext(original_filename)[1].upper().replace(".", ""),
    }
    merged_supp = _merge_supplement_from_plaintext(full_text)
    parsed.update(merged_supp)
    mil_merge = _merge_military_extractions(
        merged_supp.get(K_MILITARY_ROWS) or [],
        _military_rows_from_docx_tables(all_tables),
        _military_rows_from_fulltext_patterns(full_text),
    )
    if mil_merge:
        parsed[K_MILITARY_ROWS] = mil_merge
    qual_tbl = _qualification_rows_from_docx_table(qual_table_rows)
    if qual_tbl:
        parsed[K_QUALIFICATION_ROWS] = qual_tbl
    edu_tbl = _education_rows_from_docx_table(edu_rows)
    parsed[K_EDUCATION_ROWS] = edu_tbl if edu_tbl else []
    exp_tbl = _experience_rows_from_career_entries(career_entries)
    parsed[K_EXPERIENCE_ROWS] = (
        exp_tbl if exp_tbl else _experience_rows_from_merged_lines(career_period, career_company, career_job)
    )
    if _should_use_ai():
        ai = _ai_extract_fields(clean_flat, original_filename, "DOCX")
        parsed = _merge_ai_result(parsed, ai, clean_flat)
    return _finalize_parsed_record(parsed)


# ─── PDF 파싱 (PyMuPDF) ─────────────────────────────────────────────────────

def extract_from_pdf(doc: fitz.Document, original_filename: str) -> dict:
    """LibreOffice 변환 PDF 또는 원본 PDF 파싱"""
    full_text = ""
    for page in doc:
        blocks = page.get_text("blocks")
        # y좌표 15px 단위로 묶어 같은 행 처리 → x 기준 정렬 (우분투 레이아웃 대응)
        blocks.sort(key=lambda b: (round(b[1] / 15) * 15, b[0]))
        for b in blocks:
            full_text += b[4].replace("\r\n", "\n").replace("\r", "\n") + "\n"

    normalized_text = _normalize_pdf_text(full_text)
    clean_text = re.sub(r"[ \t]+", " ", normalized_text)
    info = _extract_basic(clean_text)
    all_dates = _DATE_RANGE.findall(clean_text)

    final_edu, edu_date = _parse_edu_from_text(clean_text, all_dates)

    career_period, career_company, career_job = _parse_career_from_text(
        clean_text, all_dates, edu_date
    )

    # PDF 표 분해 포맷 보강 (평문 학력에 전공 없고 표 파싱에는 있으면표 쪽 우선)
    pdf_edu, pdf_career_period, pdf_career_company, pdf_career_job = _parse_pdf_sections(normalized_text)
    if pdf_edu:
        if _is_bad_edu_value(final_edu):
            final_edu = pdf_edu
        elif (
            not _merged_edu_blob_has_department(final_edu)
            and _merged_edu_blob_has_department(pdf_edu)
        ):
            final_edu = pdf_edu
    if not career_period and pdf_career_period:
        career_period = pdf_career_period
    if not career_company and pdf_career_company:
        career_company = pdf_career_company
    if not career_job and pdf_career_job:
        career_job = pdf_career_job

    parsed = {
        **info,
        K_FINAL_EDU: final_edu,
        K_CAREER_PD: career_period,
        K_CAREER_CO: career_company,
        K_CAREER_ROLE: career_job,
        K_FILE_TYPE: os.path.splitext(original_filename)[1].upper().replace(".", ""),
    }
    merged_supp = _merge_supplement_from_plaintext(clean_text)
    parsed.update(merged_supp)
    mil_pdf = _merge_military_extractions(
        merged_supp.get(K_MILITARY_ROWS) or [],
        [],
        _military_rows_from_fulltext_patterns(normalized_text),
    )
    if mil_pdf:
        parsed[K_MILITARY_ROWS] = mil_pdf
    parsed[K_EDUCATION_ROWS] = []
    parsed[K_EXPERIENCE_ROWS] = _experience_rows_from_merged_lines(career_period, career_company, career_job)
    if _should_use_ai():
        ai = _ai_extract_fields(clean_text, original_filename, "PDF")
        parsed = _merge_ai_result(parsed, ai, clean_text)
    return _finalize_parsed_record(parsed)


def extract_from_hwp_pyhwp(filepath: str, original_filename: str) -> dict:
    """pyhwp(hwp5)로 HWP 텍스트를 직접 추출해 파싱"""
    try:
        from hwp5.filestructure import Hwp5File
        from hwp5.recordstream import read_records
    except ImportError:
        subprocess.run(["pip", "install", "pyhwp", "-q"], check=False)
        from hwp5.filestructure import Hwp5File
        from hwp5.recordstream import read_records

    hwp_doc = Hwp5File(filepath)
    text_parts: list[str] = []

    for section_name in hwp_doc.text:
        section_stream = hwp_doc.text[section_name].open()
        for record in read_records(section_stream):
            # 67 = HWPTAG_PARA_TEXT
            if record.get("tagid") != 67:
                continue
            payload = record.get("payload", b"")
            if not payload:
                continue
            decoded = payload.decode("utf-16-le", errors="ignore")
            # HWP 제어문자 제거 (줄바꿈은 유지)
            cleaned = "".join(
                ch if (ch >= " " or ch in "\n\r\t") else " "
                for ch in decoded
            )
            cleaned = re.sub(r"[ \t]+", " ", cleaned).strip()
            if cleaned:
                text_parts.append(cleaned)

    text = "\n".join(text_parts)
    if not text.strip():
        raise ValueError("pyhwp 추출 결과가 비어 있습니다.")

    normalized_text = _normalize_hwp_table_text(text)
    clean_text = re.sub(r"[ \t]+", " ", normalized_text)
    info = _extract_basic(clean_text)
    all_dates = _DATE_RANGE.findall(clean_text)

    final_edu, edu_date = _parse_edu_from_text(clean_text, all_dates)
    career_period, career_company, career_job = _parse_career_from_text(
        clean_text, all_dates, edu_date
    )

    # HWP 표 분해 포맷 보강
    hwp_edu, hwp_career_period, hwp_career_company, hwp_career_job = _parse_hwp_sections(normalized_text)
    if hwp_edu:
        if _is_bad_edu_value(final_edu):
            final_edu = hwp_edu
        elif (
            not _merged_edu_blob_has_department(final_edu)
            and _merged_edu_blob_has_department(hwp_edu)
        ):
            final_edu = hwp_edu
    if not career_period and hwp_career_period:
        career_period = hwp_career_period
    if not career_company and hwp_career_company:
        career_company = hwp_career_company
    if not career_job and hwp_career_job:
        career_job = hwp_career_job

    parsed = {
        **info,
        K_FINAL_EDU: final_edu,
        K_CAREER_PD: career_period,
        K_CAREER_CO: career_company,
        K_CAREER_ROLE: career_job,
        K_FILE_TYPE: os.path.splitext(original_filename)[1].upper().replace(".", ""),
    }
    merged_supp = _merge_supplement_from_plaintext(clean_text)
    parsed.update(merged_supp)
    mil_hwp = _merge_military_extractions(
        merged_supp.get(K_MILITARY_ROWS) or [],
        [],
        _military_rows_from_fulltext_patterns(normalized_text),
    )
    if mil_hwp:
        parsed[K_MILITARY_ROWS] = mil_hwp
    parsed[K_EDUCATION_ROWS] = []
    parsed[K_EXPERIENCE_ROWS] = _experience_rows_from_merged_lines(career_period, career_company, career_job)
    if _should_use_ai():
        ai = _ai_extract_fields(clean_text, original_filename, "HWP")
        parsed = _merge_ai_result(parsed, ai, clean_text)
    return _finalize_parsed_record(parsed)


# ─── 엑셀 저장 ──────────────────────────────────────────────────────────────

_EXCEL_COLUMNS = [
    K_ID,
    K_NAME, K_BIRTH_DATE, K_CONTACT, K_EMAIL, K_ADDR,
    K_ORIGINAL_JOB_ROLE, K_FINAL_EDU,
    K_CAREER_PD, K_CAREER_CO, K_CAREER_ROLE,
    K_FILE_TYPE,
]


_EXCEL_COL_WIDTH_MIN = 12.0
_EXCEL_COL_WIDTH_MAX = 56.0
_EXCEL_ROW_LINE_HEIGHT = 17.0
_EXCEL_HEADER_ROW_HEIGHT = 22.0


def _excel_display_units(value: object) -> float:
    """열 폭 추정용: 한글 등은 대략 ASCII 대비 넓게 잡음."""
    if value is None:
        return 0.0
    w = 0.0
    for ch in str(value):
        w += 2.0 if ord(ch) > 127 else 1.0
    return w


def _wrapped_line_count(text: str, column_width_units: float) -> int:
    """줄바꿈·열 폭 기준으로 필요한 행 수(대략)."""
    if not text:
        return 1
    # 열 너비(글자 단위)에 비례해 한 줄에 들어가는 표시 길이 상한
    usable = max(8.0, (column_width_units or _EXCEL_COL_WIDTH_MIN) * 1.5)
    total = 0
    for block in str(text).split("\n"):
        du = _excel_display_units(block)
        total += max(1, math.ceil(du / usable))
    return max(1, total)


def _apply_wrap_text(path: str) -> None:
    """저장된 엑셀에 열 너비·줄바꿈·행 높이를 맞춰 글자가 잘리지 않게 함."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb.active

    if ws.max_row < 1 or ws.max_column < 1:
        wb.save(path)
        return

    # 1) 열별 실제 문자(표시 길이) 기준 최소 너비
    widths: dict[int, float] = {}
    for col_idx in range(1, ws.max_column + 1):
        max_u = 0.0
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            for cell in row:
                max_u = max(max_u, _excel_display_units(cell.value))
        w = max(_EXCEL_COL_WIDTH_MIN, min(_EXCEL_COL_WIDTH_MAX, max_u / 1.85 + 4.0))
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = w
        widths[col_idx] = w

    # 헤더 행
    align_header = Alignment(wrap_text=True, vertical="center")
    top_align = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.alignment = align_header
    ws.row_dimensions[1].height = _EXCEL_HEADER_ROW_HEIGHT

    # 데이터: 모든 셀에 wrap — 열폭 한도 초과 문장은 셀 안에서 줄바꿈되어 잘리지 않게 함
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = top_align

    # 행 높이: 각 셀의 논리적 줄 수(개행 + 열폭 초과 줄바꿈) 반영
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = 1
        for cell in row:
            if not cell.value:
                continue
            cw = widths.get(cell.column, _EXCEL_COL_WIDTH_MIN)
            max_lines = max(max_lines, _wrapped_line_count(str(cell.value), cw))
        r = row[0].row
        ws.row_dimensions[r].height = max_lines * _EXCEL_ROW_LINE_HEIGHT

    wb.save(path)


def save_to_excel(data_list: list[dict]) -> None:
    df_new = pd.DataFrame(data_list)

    # 누락된 컬럼 빈 문자열로 채우고 지정 순서로 정렬
    for col in _EXCEL_COLUMNS:
        if col not in df_new.columns:
            df_new[col] = ""
    # 이미지 양식 고정: 지정 컬럼만 내보냄(추가 컬럼 제거)
    df_new = df_new[_EXCEL_COLUMNS]

    if not os.path.exists(OUTPUT_PATH):
        df_new.to_excel(OUTPUT_PATH, index=False, engine="openpyxl")
    else:
        df_old = pd.read_excel(OUTPUT_PATH, engine="openpyxl")
        # 기존 파일에도 새 컬럼 추가
        for col in _EXCEL_COLUMNS:
            if col not in df_old.columns:
                df_old[col] = ""
        df_final = (
            pd.concat([df_old, df_new], ignore_index=True)
            .drop_duplicates(subset=[K_NAME, K_CONTACT], keep="last")
        )
        # 이미지 양식 고정: 지정 컬럼만 유지
        df_final = df_final[_EXCEL_COLUMNS]
        df_final.to_excel(OUTPUT_PATH, index=False, engine="openpyxl")

    _apply_wrap_text(OUTPUT_PATH)


def export_data_list_to_excel_path(data_list: list[dict], path: str) -> None:
    """감지된 항목만 엑셀 파일(path)에 저장(기존 백엔드 `save_to_excel`은 폴더 누적용)."""
    if not data_list:
        pd.DataFrame(columns=_EXCEL_COLUMNS).to_excel(path, index=False, engine="openpyxl")
        return
    df_new = pd.DataFrame(data_list)
    for col in _EXCEL_COLUMNS:
        if col not in df_new.columns:
            df_new[col] = ""
    df_new = df_new[_EXCEL_COLUMNS]
    df_new.to_excel(path, index=False, engine="openpyxl")
    _apply_wrap_text(path)


def parse_resume_file(file_path: str) -> dict:
    """단일 이력서 파일을 파싱해 필드 dict를 반환합니다(API·배치 공통)."""
    original_name = os.path.basename(file_path)
    if original_name == OUTPUT_FILENAME or "temp_conv_" in original_name:
        raise ValueError("이 파일은 처리할 수 없습니다.")
    ext = os.path.splitext(original_name)[1].lower()
    if ext == ".docx":
        return extract_from_docx(file_path, original_name)
    if ext == ".hwp":
        try:
            return extract_from_hwp_pyhwp(file_path, original_name)
        except Exception as e:
            print(f"[WARN] {original_name} pyhwp 직접 파싱 실패, PDF 변환으로 폴백: {e}")
    current_pdf: str | None = None
    is_temporary = False
    if ext == ".pdf":
        current_pdf = file_path
    else:
        temp_name = f"temp_conv_{os.path.splitext(original_name)[0]}.pdf"
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", BASE_DIR, file_path],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        gen_pdf = os.path.join(BASE_DIR, os.path.splitext(original_name)[0] + ".pdf")
        current_pdf = os.path.join(BASE_DIR, temp_name)
        if os.path.exists(gen_pdf):
            os.rename(gen_pdf, current_pdf)
            is_temporary = True
    if not (current_pdf and os.path.exists(current_pdf)):
        raise FileNotFoundError("PDF로 변환된 파일을 찾을 수 없습니다.")
    try:
        with fitz.open(current_pdf) as pdf_doc:
            return extract_from_pdf(pdf_doc, original_name)
    finally:
        if is_temporary and current_pdf and os.path.exists(current_pdf):
            try:
                os.remove(current_pdf)
            except OSError:
                pass


# ─── 메인 ───────────────────────────────────────────────────────────────────

def main() -> None:
    target_files: list[str] = []
    for ext in ("*.docx", "*.doc", "*.hwp", "*.pptx", "*.pdf"):
        target_files.extend(glob.glob(os.path.join(BASE_DIR, ext)))
    if not target_files:
        return

    all_parsed: list[dict] = []

    for file_path in target_files:
        original_name = os.path.basename(file_path)
        if original_name == OUTPUT_FILENAME or "temp_conv_" in original_name:
            continue
        try:
            all_parsed.append(parse_resume_file(file_path))
        except Exception as e:
            print(f"[WARN] {original_name} 처리 실패: {e}")

    if all_parsed:
        save_to_excel(all_parsed)
        print(f"완료: {len(all_parsed)}건 추출 → {OUTPUT_FILENAME}")


if __name__ == "__main__":
    main()
